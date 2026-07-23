from typing import override, Any
from django.http import response, HttpResponseRedirect, HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.views.generic import ListView, CreateView, DeleteView, UpdateView, View, FormView
from django.core.paginator import Paginator
from django.conf import settings
from bolsa.models import Aspirante
from .models import CAlta, CBaja, TMovimiento
from .forms import CAltaForm
from strorganizativa.models import Departamento, CargoPlantilla
from nomencladores.models import NSalario, NCausaAltaBaja, NRol, NProvincia, NTipoContrato
from django.urls import reverse_lazy, reverse
from configuracion.models import Configuracion
from django.template.loader import get_template
from django.db.models import Q, ProtectedError, Value
from django.db.models.functions import Concat
from django.db import transaction
from xhtml2pdf import pisa
from datetime import datetime, timedelta
from .forms import CAltaForm, MovimientoForm
from docxtpl import DocxTemplate, RichText
from io import BytesIO
import os
import traceback
import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, Http404
from datetime import datetime
from .forms import ExportarContratoWordForm



CAT_MAP = {
    'TEC': 'T',
    'ADM': 'A',
    'SER': 'S',
    'OPE': 'O',
    'CDI': 'C',
    'CEJ': 'C',
}
#?ONTRATO
class ContratoListView(ListView):
    model = CAlta
    template_name = "pages/contrato/list_contrato.html"
    def get_paginate_by(self, queryset):
        # PROTECCIÓN: Si viene vacío (''), usar 8 por defecto
        page_size = self.request.GET.get('per_page')
        return int(page_size) if page_size else 8
        
    ordering = ['-fecha_alta']  
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = CAltaForm()
        context['search_url'] = 'search_contrato'
        context['causas_baja'] = NCausaAltaBaja.objects.filter(alta=False)
        context['provincias_list'] = NProvincia.objects.all()
        context['niveles_educ'] = Aspirante._meta.get_field('nivel_educ').choices

        # LA CLAVE: Enviar la variable al HTML desde la carga inicial
        page_size = self.request.GET.get('per_page')
        context['current_page_size'] = str(page_size) if page_size else '8'
        return context
    

class MovimientoNominaListView(ListView):
    model = CAlta
    # Usaremos la nueva ruta de plantilla que acordamos
    template_name = "pages/movimientos/list_movimientos.html" 
    paginate_by = 8
    ordering = ['-fecha_alta']

    def get_queryset(self):
        # FILTRO CLAVE: Solo mostramos los marcados como "En Proceso"
        return CAlta.objects.filter(en_proceso_movimiento=True).order_by('-fecha_alta')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pasamos search_url por si quieres implementar búsqueda aquí también en el futuro
        # (Por ahora reutilizaremos lógica básica o lo dejaremos simple)
        context['titulo'] = "Movimientos de Nómina Pendientes"
        return context


# --- Función de apoyo para entender los Grupos Escala ---
def nivel_romano_a_int(romano):
    mapa = {'I': 1, 'V': 5, 'X': 10, 'L': 50}
    res, prev = 0, 0
    for c in reversed(romano.upper()):
        val = mapa.get(c, 0)
        if val < prev: res -= val
        else: res += val
        prev = val
    return res


def cargar_roles(request):
    cargo_id = request.GET.get('cargo')
    roles_disponibles = []
    
    if cargo_id:
        try:
            cargo = CargoPlantilla.objects.select_related('ncargo').get(pk=cargo_id)
            grupo = cargo.ncargo.grupo_escala
            categoria = cargo.ncargo.cat_ocupacional
            nivel_num = nivel_romano_a_int(grupo.nivel) # Leemos el número del grupo
            
            # 1. LA LIMPIEZA (Tu código original): Buscar en NSalario qué roles tienen dinero (> 0)
            salarios_con_fondo = NSalario.objects.filter(
                grupo_escala=grupo, 
                monto__gt=0
            ).values_list('rol_id', flat=True).distinct()
            
            # Convertimos a Set para manipular fácil
            ids_validos = set(salarios_con_fondo)
            
            # Buscamos los IDs de los roles clave
            rol_cuadro_obj = NRol.objects.filter(tipo__iexact="Cuadro").first()
            rol_decisorio_obj = NRol.objects.filter(tipo__iexact="Decisorio").first()
            
            id_cuadro = rol_cuadro_obj.id if rol_cuadro_obj else None
            id_decisorio = rol_decisorio_obj.id if rol_decisorio_obj else None
            
            # 2. LA CATEGORÍA (Tu código original)
            if id_cuadro:
                # Si ES directivo, nos aseguramos que vea la opción "Cuadro"
                if categoria in ['CDI', 'CEJ']:
                    ids_validos.add(id_cuadro)
                else:
                    if id_cuadro in ids_validos:
                        ids_validos.remove(id_cuadro)

            # 3. EL PORTERO INTELIGENTE: Reglas estrictas de los Grupos Escala
            
            # Regla Grupos I al XIX: PROHIBIDO Cuadro (Incluso si es CDI/CEJ lo quitamos)
            if nivel_num <= 19:
                if id_cuadro and id_cuadro in ids_validos:
                    ids_validos.remove(id_cuadro)
                    
            # Regla Grupos XX y XXI: Permite todos (No quitamos nada)
            elif 20 <= nivel_num <= 21:
                pass
                
            # Regla Grupos XXII al XXIV: SOLO Cuadro o Decisorio (Destruimos Apoyo y Fundamental)
            elif 22 <= nivel_num <= 24:
                permitidos = set()
                if id_cuadro and id_cuadro in ids_validos: 
                    permitidos.add(id_cuadro)
                if id_decisorio and id_decisorio in ids_validos: 
                    permitidos.add(id_decisorio)
                ids_validos = permitidos 
                
            # Regla Grupo XXV: SOLO Cuadro
            elif nivel_num == 25:
                permitidos = set()
                if id_cuadro and id_cuadro in ids_validos: 
                    permitidos.add(id_cuadro)
                ids_validos = permitidos

            # 4. Query Final
            if ids_validos:
                roles_disponibles = NRol.objects.filter(id__in=ids_validos).order_by('tipo')
            else:
                roles_disponibles = NRol.objects.none()

        except Exception as e:
            print(f"Error cargando roles: {e}")
            pass
            
    return render(request, 'pages/contrato/partials/options_roles.html', {'roles': roles_disponibles})

def obtener_datos_previos(request):
    """
    Busca datos previos en CAlta (Activos) O CBaja (Histórico de Bajas)
    para recuperar No. Expediente y Registro Militar.
    """
    aspirante_id = request.GET.get('aspirante_id')
    if not aspirante_id:
        return JsonResponse({'existe': False})

    # 1. Buscar primero en Bajas (lo más probable si se está recontratando)
    baja_reciente = CBaja.objects.filter(aspirante_id=aspirante_id).order_by('-fecha_baja').first()
    
    if baja_reciente:
        return JsonResponse({
            'existe': True,
            'no_expediente': baja_reciente.no_expediente,
            'reg_militar': baja_reciente.reg_militar
        })

    # 2. Si no está en bajas, buscar en Activos (casos raros de pluriempleo o errores)
    alta_reciente = CAlta.objects.filter(aspirante_id=aspirante_id).order_by('-fecha_alta').first()
    
    if alta_reciente:
        return JsonResponse({
            'existe': True,
            'no_expediente': alta_reciente.no_expediente,
            'reg_militar': alta_reciente.reg_militar
        })

    # 3. No se encontró nada
    return JsonResponse({'existe': False})

def search_contratos(request):
    # 1. Obtener parámetros de filtros
    query = request.GET.get('filter_contrato', '').strip()
    page_num = request.GET.get('page', 1)
    page_size = request.GET.get('per_page')
    if not page_size:
        page_size = '8'
    
    # Filtros del Aspirante asociado
    provincia_id = request.GET.get('provincia')
    municipio_id = request.GET.get('municipio')
    nivel_educ = request.GET.get('nivel_educ')
    especialidad_id = request.GET.get('especialidad')
    sexo = request.GET.get('sexo')
    raza = request.GET.get('raza')
    grado_cientifico = request.GET.get('grado_cientifico')

    # 2. QuerySet Base con el Grafo Relacional expandido (Corrección N+1)
    qs = CAlta.objects.select_related(
        'aspirante', 
        'cargo',
        'cargo__departamento',
        'cargo__ncargo',
        'cargo__ncargo__grupo_escala',
        'cargo__rol',
        'rol',
        'tridente'
    ).all().order_by('-fecha_alta')

    # 3. Aplicar Filtros "Embudo" (Relación aspirante__)
    if provincia_id:
        qs = qs.filter(aspirante__provincia_id=provincia_id)
    if municipio_id:
        qs = qs.filter(aspirante__municipio_id=municipio_id)
    if nivel_educ:
        qs = qs.filter(aspirante__nivel_educ=nivel_educ)
    if especialidad_id:
        qs = qs.filter(aspirante__especialidad_id=especialidad_id)
    if sexo:
        qs = qs.filter(aspirante__sexo=sexo)
    if raza:
        qs = qs.filter(aspirante__raza=raza)
    if grado_cientifico:
        qs = qs.filter(aspirante__grado_cientifico=grado_cientifico)

    # 4. Buscador Inteligente (Concatenación)
    if query:
        qs = qs.annotate(
            nombre_completo=Concat(
                'aspirante__nombre', Value(' '), 
                'aspirante__papellido', Value(' '), 
                'aspirante__sapellido'
            )
        ).filter(
            Q(no_expediente__icontains=query) |
            Q(nombre_completo__icontains=query)|
            Q(aspirante__doc_identidad__icontains=query)
        )

    # 5. ORDENAMIENTO COMPLETO EN EL SERVIDOR (Todas las páginas)
    sort_col = request.GET.get('sort', '')
    order = request.GET.get('order', 'asc')
    
    # Convertimos el QuerySet a una Lista. Aquí se ejecuta la única consulta SQL.
    contratos_list = list(qs)
    
    if sort_col:
        reverse_sort = (order == 'desc')
        
        if sort_col == 'expediente':
            contratos_list.sort(key=lambda x: int(x.no_expediente) if x.no_expediente.isdigit() else 0, reverse=reverse_sort)
        elif sort_col == 'nombre':
            contratos_list.sort(key=lambda x: f"{x.aspirante.nombre} {x.aspirante.papellido} {x.aspirante.sapellido}".lower(), reverse=reverse_sort)
        elif sort_col == 'area':
            contratos_list.sort(key=lambda x: x.cargo.departamento.descripcion.lower() if x.cargo and x.cargo.departamento else "", reverse=reverse_sort)
        elif sort_col == 'cat':
            contratos_list.sort(key=lambda x: getattr(x.cargo.ncargo, 'get_cat_ocupacional_display')() if x.cargo and hasattr(x.cargo, 'ncargo') else "", reverse=reverse_sort)
        elif sort_col == 'grupo':
            def get_grupo_val(c):
                nivel = c.cargo.ncargo.grupo_escala.nivel if c.cargo and c.cargo.ncargo and hasattr(c.cargo.ncargo, 'grupo_escala') and c.cargo.ncargo.grupo_escala else ""
                return nivel_romano_a_int(nivel)
            contratos_list.sort(key=get_grupo_val, reverse=reverse_sort)
        elif sort_col == 'rol':
            jerarquia = {"APOYO": 1, "FUNDAMENTAL": 2, "DECISORIO": 3, "CUADRO": 4}
            def get_rol_val(c):
                r = c.rol or (c.cargo.rol if c.cargo else None)
                rol_str = r.tipo.upper() if r else "CUADRO"
                return jerarquia.get(rol_str, 0)
            contratos_list.sort(key=get_rol_val, reverse=reverse_sort)
        elif sort_col == 'tridente':
            contratos_list.sort(key=lambda x: x.tridente.tipo if x.tridente else "", reverse=reverse_sort)
        elif sort_col == 'salario':
            contratos_list.sort(key=lambda x: x.calcular_salario_escala() or 0, reverse=reverse_sort)

    # 6. Paginación SOBRE la lista ya ordenada
    paginator = Paginator(contratos_list, int(page_size))
    page_obj = paginator.get_page(page_num)

    return render(request, 'pages/contrato/partials/filter_contratos_list.html', {
        'object_list': page_obj,  
        'page_obj': page_obj,     
        'paginator': paginator,   
        'search_url': 'search_contrato', 
        'current_page_size': str(page_size)  
    })




def validar_datos_contrato(request):
    expediente = request.GET.get('no_expediente', None)
    data = {
        'expediente_existe': False
    }
    if expediente:
        # Verifica si existe algún contrato con ese No. Expediente
        if CAlta.objects.filter(no_expediente=expediente).exists():
            data['expediente_existe'] = True
            
    return JsonResponse(data)

def solicitar_movimiento_nomina(request, pk):
    """
    Alterna el estado de 'En Proceso de Movimiento' de un contrato.
    """
    if request.method == "POST":
        try:
            contrato = get_object_or_404(CAlta, pk=pk)
            
            # Leemos el parámetro enviado por JS
            es_cancelacion = request.POST.get('cancelar') == 'true'
            
            if es_cancelacion:
                contrato.en_proceso_movimiento = False
                mensaje = 'Movimiento cancelado. El contrato regresó a su estado normal.'
            else:
                contrato.en_proceso_movimiento = True
                mensaje = 'Contrato enviado a la bandeja de movimientos pendientes.'
                
            contrato.save()
            return JsonResponse({'success': True, 'message': mensaje})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

# orbith/contratos/views.py

def validar_plazas_cargo(request):
    cargo_id = request.GET.get('cargo_id', None)
    tipo_id = request.GET.get('tipo_id', None)
    data = {
        'plazas_agotadas': False,
        'mensaje': ''
    }
    
    if cargo_id:
        try:
            cargo = CargoPlantilla.objects.select_related('ncargo').get(pk=cargo_id)
            
            # 1. Evaluamos la regla de negocio del Tipo de Contrato
            if tipo_id:
                try:
                    tipo_contrato = NTipoContrato.objects.get(pk=tipo_id)
                    # Si el contrato NO ocupa plaza (Ej: Adiestrado, Determinado), damos luz verde directa
                    if not tipo_contrato.ocupa_plaza:
                        return JsonResponse(data)
                except NTipoContrato.DoesNotExist:
                    pass
        
            if cargo.cant_cubierta_real >= cargo.cant_aprobada:
                data['plazas_agotadas'] = True
                data['mensaje'] = f"El Cargo '{cargo.ncargo.descripcion}' no tiene plazas oficiales disponibles ({cargo.cant_cubierta_real}/{cargo.cant_aprobada})"
                
        except CargoPlantilla.DoesNotExist:
            pass
            
    return JsonResponse(data)
            
    

def cargar_departamentos(request):
    unidad_id = request.GET.get('unidad')
    departamentos = Departamento.objects.none()
    
    if unidad_id:
        try:
            departamentos = Departamento.objects.filter(unidad_organizativa_id=unidad_id).order_by('descripcion')
        except (ValueError, TypeError):
            pass

    return render(request, 'pages/contrato/partials/options_generico.html', {'opciones': departamentos})

def cargar_cargos(request):
    dpto_id = request.GET.get('departamento')
    cargos = CargoPlantilla.objects.none()
    
    if dpto_id:
        try:
            # Filtramos cargos del departamento
            cargos = CargoPlantilla.objects.filter(departamento_id=dpto_id).select_related('ncargo').order_by('ncargo__descripcion')
        except (ValueError, TypeError):
            pass
            
    return render(request, 'pages/contrato/partials/options_cargos.html', {'cargos': cargos})



def historico_trabajador(request, aspirante_id):

    try:
        aspirante = Aspirante.objects.get(pk=aspirante_id)
        
        # 1. Buscamos primero en contratos activos
        contrato_data = CAlta.objects.filter(aspirante_id=aspirante_id).order_by('-fecha_alta').first()
        
        # 2. Si no hay contrato activo (está de baja), buscamos en el historial de bajas para recuperar el expediente vitalicio
        if not contrato_data:
            contrato_data = CBaja.objects.filter(aspirante_id=aspirante_id).order_by('-fecha_alta').first()
            
        exp_num = contrato_data.no_expediente if contrato_data else "---"
        encabezado = {
            'nombre_completo': f"{aspirante.nombre} {aspirante.papellido} {aspirante.sapellido}",
            'ci': aspirante.doc_identidad,
            'id': aspirante.id, 
            'expediente': exp_num
        }
    except Aspirante.DoesNotExist:
        return JsonResponse({'data': [], 'encabezado': {'nombre_completo': 'Desconocido', 'ci': ''}})

    def fmt_dinero(valor):
        try: return f"{float(valor):.2f}"
        except: return "0.00"

    hitos = []

    # --- FUNCIÓN AUXILIAR PARA BUSCAR EL ID DEL MOVIMIENTO ---
    def obtener_id_movimiento(aspirante_id, expediente, tipo_busqueda):
        # Buscamos en TMovimiento el registro que coincide con este expediente y tipo
        mov = TMovimiento.objects.filter(
            aspirante_id=aspirante_id, 
            no_expediente=expediente, 
            tipo_movimiento=tipo_busqueda
        ).first()
        return mov.pk if mov else None

    # --- FUENTE A: Altas Activas ---
    altas_activas = CAlta.objects.filter(aspirante_id=aspirante_id)
    for a in altas_activas:
        # CORRECCIÓN: Ignorar el movimiento fantasma de "Alta Inicial" para buscar el pasado real
        mov_id = obtener_id_movimiento(aspirante_id, a.no_expediente, 'Alta Inicial')
        primer_mov = TMovimiento.objects.filter(contrato=a).exclude(tipo_movimiento='Alta Inicial').order_by('fecha_efectiva').first()
        
        if primer_mov and primer_mov.cargo_anterior != "---":
            cargo_inicial = primer_mov.cargo_anterior
            unidad_inicial = primer_mov.unidad_anterior or "---"
            salario_inicial = primer_mov.salario_anterior
        else:
            cargo_inicial = a.cargo.ncargo.descripcion if a.cargo else "---"
            unidad_inicial = a.cargo.departamento.unidad_organizativa.descripcion if (a.cargo and a.cargo.departamento) else "---"
            salario_inicial = a.calcular_salario_escala()

        hitos.append({
            'id': mov_id,
            'fecha_orden': a.fecha_alta,
            'fecha_inicio': a.fecha_alta,
            'evento': 'ALTA_PENDIENTE', 
            'expediente': a.no_expediente,
            'unidad': unidad_inicial,
            'cargo': cargo_inicial,
            'salario': fmt_dinero(salario_inicial),
            'es_baja': False
        })

    # --- FUENTE B: Altas de Contratos Cerrados ---
    altas_viejas = CBaja.objects.filter(aspirante_id=aspirante_id)
    for b in altas_viejas:
        if b.fecha_alta:
            mov_id_alta = obtener_id_movimiento(aspirante_id, b.no_expediente, 'Alta Inicial')
            mov_id_baja = obtener_id_movimiento(aspirante_id, b.no_expediente, 'Baja')
            # CORRECCIÓN: Igual que arriba, ignoramos "Alta Inicial"
            primer_mov_baja = TMovimiento.objects.filter(
                aspirante_id=aspirante_id, 
                no_expediente=b.no_expediente
            ).exclude(tipo_movimiento='Alta Inicial').order_by('fecha_efectiva').first()

            if primer_mov_baja and primer_mov_baja.cargo_anterior != "---":
                cargo_baja_ini = primer_mov_baja.cargo_anterior
                unidad_baja_ini = primer_mov_baja.unidad_anterior or "---"
                salario_baja_ini = primer_mov_baja.salario_anterior
            else:
                cargo_baja_ini = b.cargo.ncargo.descripcion if b.cargo else "---"
                unidad_baja_ini = b.cargo.departamento.unidad_organizativa.descripcion if (b.cargo and b.cargo.departamento) else "---"
                salario_baja_ini = b.cargo.ncargo.salario_basico if b.cargo else 0

            hitos.append({
                'id': mov_id_alta,
                'fecha_orden': b.fecha_alta,
                'fecha_inicio': b.fecha_alta,
                'evento': 'ALTA_PENDIENTE',
                'expediente': b.no_expediente,
                'unidad': unidad_baja_ini,
                'cargo': cargo_baja_ini,
                'salario': fmt_dinero(salario_baja_ini),
                'es_baja': False
            })
            
            # El evento de BAJA explícito
            hitos.append({
                'id': mov_id_baja,
                'fecha_orden': b.fecha_baja,
                'fecha_inicio': b.fecha_baja,
                'evento': 'Baja',
                'expediente': b.no_expediente,
                'unidad': "---",
                'cargo': "---",
                'salario': "---",
                'es_baja': True
            })

    # --- FUENTE C: Movimientos ---
    # CORRECCIÓN: Filtramos los movimientos que crea el sistema por detrás ('Baja' y 'Alta Inicial')
    movimientos = TMovimiento.objects.filter(aspirante_id=aspirante_id).exclude(tipo_movimiento__in=['Baja', 'Alta Inicial'])
    
    for m in movimientos:
        tipo = "Movimiento Salario"
        if m.unidad_anterior != m.unidad_nueva or m.cargo_anterior != m.cargo_nuevo:
            tipo = "Movimiento Cargo"

        hitos.append({
            'id': m.pk,
            'fecha_orden': m.fecha_efectiva,
            'fecha_inicio': m.fecha_efectiva,
            'evento': tipo,
            'expediente': m.no_expediente,
            'unidad': m.unidad_nueva or "---",
            'cargo': m.cargo_nuevo,
            'salario': fmt_dinero(m.salario_nuevo),
            'es_baja': False
        })

    # --- ORDENACIÓN Y LÓGICA CRONOLÓGICA ---
    hitos.sort(key=lambda x: x['fecha_orden'])

    lista_final = []
    total = len(hitos)
    primer_alta = True
    
    # LA CLAVE: Leer el estado real del trabajador para saber si el último ciclo sigue vivo
    esta_activo_ahora = (aspirante.estado == 'ACTIVO')

    for i in range(total):
        item = hitos[i]
        
        # 1. Definir si es Alta Inicial o Recontratación
        if item['evento'] == 'ALTA_PENDIENTE':
            if primer_alta:
                item['evento'] = 'Alta Inicial'
                primer_alta = False
            else:
                item['evento'] = 'Recontratación'

        # =======================================================
        # 2. CALCULAR FECHA FIN (SOLUCIÓN DEFINITIVA)
        # =======================================================
        if i == total - 1:
            # Es el último evento de la historia:
            if item['es_baja']:
                # Si lo último registrado es una Baja, su estado actual es "En Baja"
                # El color será rojo porque el evento es 'Baja' (text-danger)
                item['fecha_fin'] = "Actualidad"
            else:
                # Si lo último es un contrato, solo es Actualidad si el aspirante está ACTIVO
                item['fecha_fin'] = "Actualidad" if esta_activo_ahora else "Cerrado"
        else:
            proximo = hitos[i+1]
            
            # REGLA A: Contrato (Alta/Mov) seguido por su propia Baja.
            # El contrato termina el mismo día que se emite la baja.
            if not item['es_baja'] and proximo['es_baja'] and item['expediente'] == proximo['expediente']:
                item['fecha_fin'] = proximo['fecha_inicio'].strftime('%d/%m/%Y')
            
            # REGLA B: Baja seguida por Recontratación (o Movimiento seguido por otro Movimiento).
            # El estado anterior termina un día antes de que empiece el nuevo hito.
            else:
                fin_dt = proximo['fecha_inicio'] - timedelta(days=1)
                item['fecha_fin'] = fin_dt.strftime('%d/%m/%Y')

        # 3. Asignar Colores (Solo para la fecha fin)
        colores = {
            'Alta Inicial': 'text-success',      # Verde
            'Recontratación': 'text-primary',    # Azul
            'Movimiento Salario': 'text-gold',   # Dorado
            'Movimiento Cargo': 'text-orange',   # Naranja
            'Baja': 'text-danger'                # Rojo
        }
        item['estado_clase'] = colores.get(item['evento'], 'text-muted')
        
        # Formatear fecha inicio
        item['fecha_inicio_str'] = item['fecha_inicio'].strftime('%d/%m/%Y')
        lista_final.append(item)

    return JsonResponse({'data': lista_final, 'encabezado': encabezado})

# --- 2. NUEVA VISTA PARA EXPORTAR HISTORIAL COMPLETO (PDF) ---
class ExportarHistoricoPDFView(View):
    def get(self, request, aspirante_id):
        from django.template.loader import get_template
        from xhtml2pdf import pisa
        
        aspirante = get_object_or_404(Aspirante, pk=aspirante_id)
        config = Configuracion.objects.first()
        
        # Súper truco: Llamamos a tu propia función para reciclar los datos ordenados
        respuesta = historico_trabajador(request, aspirante_id)
        datos_historial = json.loads(respuesta.content)['data']
        
        context = {
            'aspirante': aspirante,
            'config': config,
            'fecha_hoy': datetime.now(),
            'historial': datos_historial
        }
        
        template = get_template('pages/reportes/pdf_historico_trabajador.html')
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Historico_{aspirante.doc_identidad}.pdf"'
        
        pisa.CreatePDF(html, dest=response)
        return response

# --- 3. NUEVA VISTA PARA LA CÁPSULA DEL TIEMPO (SOLO LECTURA) ---
def movimiento_detalle_readonly(request, pk):
    movimiento = get_object_or_404(TMovimiento, pk=pk)
    context = {
        'object': movimiento,
        'movimiento': movimiento,
        'aspirante': movimiento.aspirante,
        'is_readonly': True,
        
        # Estas son las variables que el modal busca para la columna DERECHA
        'initial_rol': movimiento.rol_nuevo,
        'initial_grupo': movimiento.grupo_escala_nuevo,
        'initial_cat': movimiento.cat_ocupacional_nuevo,
        'initial_salario_escala': movimiento.salario_nuevo,
    }
    return render(request, 'pages/contrato/movimiento_nomina.html', context)

class ContratoCreateView(CreateView):
    model = CAlta
    form_class = CAltaForm
    template_name = "pages/contrato/add_contrato.html"
    success_url = reverse_lazy('list_aspir')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # <-- aquí pasas el usuario
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # aspirante_id viene en la URL
        aspirante_id = self.kwargs['aspirante_id']
        context['aspirante'] = get_object_or_404(Aspirante, doc_identidad=aspirante_id)
        return context

    def form_valid(self, form):
        self.object = form.save(commit=False)
        
        if self.object.cargo and self.object.cargo.rol:
            self.object.rol = self.object.cargo.rol
        
        aspirante_id = self.kwargs['aspirante_id']
        self.object.aspirante = get_object_or_404(Aspirante, doc_identidad=aspirante_id)
        
        # --- FASE 1: GUARDAR EN SESIÓN, NO EN BD ---
        # Convertimos los datos críticos del formulario a un diccionario serializable
        contrato_data = form.cleaned_data.copy()
        
        from decimal import Decimal # <--- Importación necesaria para detectar los decimales
        
        # Limpiar datos no serializables (fechas, objetos, decimales) para la sesión
        for key, value in contrato_data.items():
            if hasattr(value, 'isoformat'): # Fechas
                contrato_data[key] = value.isoformat()
            elif hasattr(value, 'pk'): # Modelos (ForeignKeys)
                contrato_data[key] = value.pk
            elif isinstance(value, Decimal): # <--- SOLUCIÓN AL ERROR: Convertir Decimales
                contrato_data[key] = str(value) # Lo guardamos como string para no perder precisión
        
        # Almacenamos el borrador
        self.request.session['contrato_borrador'] = contrato_data
        self.request.session['aspirante_borrador'] = aspirante_id
        
        # NO HACEMOS self.object.save()
        
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            from django.urls import reverse
            return JsonResponse({
                'form_is_valid': True,
                'message': 'Borrador guardado. Continuando al movimiento de nómina...',
                # En la Fase 2 crearemos esta URL para el modal del Movimiento, 
                # por ahora usaremos un endpoint temporal o el que vayamos a crear:
                'redirect_to_wizard': reverse('wizard_movimiento_nomina', kwargs={'aspirante_id': aspirante_id}) 
            })
        
        # Evitamos el guardado por defecto de CreateView
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(self.success_url)

    def form_invalid(self, form):
        # 1. Imprimir en consola para depuración
        print("🔴 ERRORES AL CREAR CONTRATO:", form.errors)

        # 2. RESPUESTA AJAX (Para evitar el crasheo)
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            
            # Capturador dinámico de errores para el SweetAlert
            error_mensaje = None
            if 'no_expediente' in form.errors:
                error_mensaje = f"El Número de Expediente ya está en uso por un contrato activo."
            elif 'cargo' in form.errors:
                error_mensaje = form.errors['cargo'][0]
            elif 'reg_militar' in form.errors:
                error_mensaje = "El campo Servicio Militar es obligatorio."
            elif form.errors:
                # Si es cualquier otro error, agarramos el primero que encuentre
                primer_campo = list(form.errors.keys())[0]
                error_mensaje = f"Verifique el campo '{primer_campo}': {form.errors[primer_campo][0]}"
            
            # Renderizamos el formulario con los errores pintados (rojo)
            html = render_to_string(
                self.template_name,
                self.get_context_data(form=form),
                request=self.request
            )
            
            # Devolvemos JSON con el HTML y el mensaje para el SweetAlert
            return JsonResponse({
                'form_is_valid': False, 
                'html_form': html,
                'error_popup': error_mensaje  # <--- Este es el dato clave
            })
            
        return super().form_invalid(form)

class ContratoUpdateView(UpdateView):
    model = CAlta
    form_class = CAltaForm
    template_name = "pages/contrato/updt_contrato.html"
    success_url = reverse_lazy('list_contrato')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        # --- TU LÓGICA ORIGINAL SE MANTIENE INTACTA ---
        context = super().get_context_data(**kwargs)

        contrato = self.object
        context['aspirante'] = contrato.aspirante

        config = Configuracion.objects.first()
        if config and config.fondo_tiempo_calc_tarif is not None:
            fondo = float(config.fondo_tiempo_calc_tarif)
        else:
            fondo = 190.6
            messages.warning(self.request, "Parámetros de configuración no encontrados. Usando valor por defecto.")
        context['fondo'] = fondo

        if config and config.porcentaje_horas_extras is not None:
            porcentaje_extra = float(config.porcentaje_horas_extras)
        else:
            porcentaje_extra = 0.25
    
        # 2) Si ya tiene cargo y tridente, calculo salario/tarifa/extras
        if contrato.cargo and contrato.tridente:
            cargo = contrato.cargo
            tridente_id = contrato.tridente.id
            grupo_escala = cargo.ncargo.grupo_escala
            rol = cargo.rol

            try:
                salario_obj = NSalario.objects.get(
                    grupo_escala=grupo_escala,
                    rol=rol,
                    tridente_id=tridente_id
                )
                monto = float(salario_obj.monto)
                
                tarifa = round(monto / fondo, 5) if fondo else 0
                extras = round((tarifa * porcentaje_extra) + tarifa, 5)
                
                context['initial_salario_escala'] = round(monto, 2)
                context['initial_tarifa_horaria'] = tarifa
                context['initial_tarifa_extras'] = extras
            except NSalario.DoesNotExist:
                context['initial_salario_escala'] = ""
                context['initial_tarifa_horaria'] = ""
                context['initial_tarifa_extras'] = ""
        else:
            context['initial_salario_escala'] = ""
            context['initial_tarifa_horaria'] = ""
            context['initial_tarifa_extras'] = ""

        return context

    # --- MÉTODOS AÑADIDOS PARA SOPORTE AJAX ---

    def form_valid(self, form):
        self.object = form.save(commit=False)
        
        if self.object.cargo and self.object.cargo.rol:
            self.object.rol = self.object.cargo.rol
            
        # --- LIMPIEZA AUTOMÁTICA DE TRIDENTE ---
        if self.object.rol and "Cuadro" in self.object.rol.tipo:
            self.object.tridente = None
            
        self.object.save()
        
        # ✅ PARA PETICIONES AJAX NORMALES (esto va ANTES del bloque HTMX)
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({
                'form_is_valid': True,
                'success_url': str(self.success_url)
            })
        
        # MAGIA HTMX: Respondemos directamente con las órdenes para el frontend
        if self.request.headers.get('HX-Request'):
            import json
            from django.http import HttpResponse
            response = HttpResponse(status=204)
            response['HX-Trigger'] = json.dumps({
                'updateContratoList': '',
                'updateCargoList': '',
                'showMessage': {'icon': 'success', 'text': 'Contrato actualizado.'},
                'closeModal': ''
            })
            return response

        messages.success(self.request, 'Contrato actualizado correctamente.')
        return super(UpdateView, self).form_valid(form)

    def form_invalid(self, form):
        # RESPUESTA AJAX (ERROR)
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string(
                self.template_name,
                self.get_context_data(form=form),
                request=self.request
            )
            return JsonResponse({'form_is_valid': False, 'html_form': html})
            
        return super().form_invalid(form)



class ContratoDeleteView(DeleteView):
    model = CAlta

    def post(self, request, *args, **kwargs):
        # Asegúrate de tener este import arriba del todo en el archivo:
        
        
        # 1. Recuperar el objeto de forma segura
        try:
            contrato: CAlta = self.get_object() # type: ignore
        except CAlta.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'El contrato que intenta eliminar no existe.'}, status=404)

        # 2. Obtener datos del formulario
        fecha_baja_str = request.POST.get('fecha_baja')
        fecha_efectiva_str = request.POST.get('fecha_efectiva')
        causa_id = request.POST.get('causa_baja')
        cobro_sistema = request.POST.get('cobro_sistema_pago') == 'true'
        completar_cargo_val = request.POST.get('completar_cargo') == 'true'
        actividad_manual = request.POST.get('actividad_realizada')
        observaciones_texto = request.POST.get('observaciones', '')


        if not fecha_baja_str or not causa_id:
            return JsonResponse({'success': False, 'message': 'Faltan datos obligatorios: Fecha de Baja o Causa.'}, status=400)

        # 3. Convertir fecha para validar
        try:
            fecha_baja = datetime.strptime(fecha_baja_str, '%Y-%m-%d').date()
            fecha_efectiva = datetime.strptime(fecha_efectiva_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
             return JsonResponse({'success': False, 'message': 'Formato de fecha inválido.'}, status=400)

        # =================================================================
        # VALIDACIÓN CRONOLÓGICA (La Barrera)
        # =================================================================
        
        
        # A. Buscar último movimiento registrado
        ultimo_mov = TMovimiento.objects.filter(contrato=contrato).order_by('-fecha_efectiva').first()
        
        # B. Definir fecha límite (El evento más reciente entre el último mov o el alta original)
        fecha_limite = ultimo_mov.fecha_efectiva if ultimo_mov else contrato.fecha_alta
        
        # C. Comparar (Baja no puede ser anterior a lo último que pasó)
        if fecha_limite and fecha_baja < fecha_limite:
            return JsonResponse({
                'success': False, 
                'message': f"No puede dar Baja con fecha {fecha_baja.strftime('%d/%m/%Y')} porque existe un evento posterior el {fecha_limite.strftime('%d/%m/%Y')}."
            }, status=400)
        # =================================================================

        try:
            # 4. Iniciar Transacción
            with transaction.atomic():
                # A. SALVAR EL HISTORIAL EXISTENTE
                TMovimiento.objects.filter(contrato=contrato).update(
                    aspirante=contrato.aspirante,
                    no_expediente=contrato.no_expediente
                )

                # B. CAPTURAR DATOS PARA EL HISTÓRICO ANTES DE BORRAR
                cargo_ant_str = contrato.cargo.ncargo.descripcion if contrato.cargo else "---"
                unidad_ant_str = contrato.cargo.departamento.unidad_organizativa.descripcion if (contrato.cargo and contrato.cargo.departamento) else "---"
                rol_ant_str = contrato.cargo.rol.tipo if (contrato.cargo and contrato.cargo.rol) else "---"
                tri_ant_str = str(contrato.tridente) if contrato.tridente else "---"
                salario_ant_val = contrato.calcular_salario_escala() or 0

                # C. CREAR EL ÚNICO MOVIMIENTO DE BAJA (Corrigiendo el error de fechas)
                # fecha_baja (Modal) = Fecha Efectiva (Cuando se va)
                # fecha_efectiva (Modal) = Fecha Documento (Cuando se firma)
                movimiento_baja = TMovimiento.objects.create(
                    aspirante=contrato.aspirante,
                    no_expediente=contrato.no_expediente,
                    contrato=None, 
                    fecha_efectiva=fecha_baja,      
                    fecha_solicitud=fecha_efectiva, 
                    tipo_movimiento="Baja",
                    observaciones=observaciones_texto, # <--- AQUÍ GUARDAMOS LA OBSERVACIÓN
                    
                    cargo_anterior=cargo_ant_str,
                    departamento_anterior=contrato.cargo.departamento.descripcion if (contrato.cargo and contrato.cargo.departamento) else "---",
                    grupo_escala_anterior=contrato.cargo.ncargo.grupo_escala.nivel if (contrato.cargo and contrato.cargo.ncargo.grupo_escala) else "---",
                    cat_ocupacional_anterior=contrato.cargo.ncargo.get_cat_ocupacional_display() if contrato.cargo else "---",
                    tipo_salario_anterior=contrato.get_tipo_salario_display() if contrato.tipo_salario else "---",
                    rol_anterior=rol_ant_str,
                    tridente_anterior=tri_ant_str,
                    unidad_anterior=unidad_ant_str,
                    salario_anterior=salario_ant_val,
                    
                    cargo_nuevo="---",
                    salario_nuevo=0,
                    unidad_nueva="---"
                )
                
                # D. Crear registro histórico en CBaja
                CBaja.objects.create(
                    aspirante=contrato.aspirante,
                    no_expediente=contrato.no_expediente,
                    tipo=contrato.tipo,
                    cargo=contrato.cargo,
                    reg_militar=contrato.reg_militar,
                    profesional=contrato.profesional,
                    fecha_baja=fecha_baja,
                    causa_baja_id=causa_id,
                    fecha_alta=contrato.fecha_alta,
                    tridente=contrato.tridente,
                    cobro_sistema_pago=cobro_sistema,
                    actividad_realizada=actividad_manual,
                    observaciones=observaciones_texto,
                    completar_cargo=completar_cargo_val
                )

                # E. Capturar el cargo antes de borrar el contrato
                cargo_obj = None
                if contrato.tipo and contrato.tipo.ocupa_plaza and contrato.cargo:
                    cargo_obj = contrato.cargo

                # Cambiar estado del aspirante a BAJA
                if contrato.aspirante:
                    contrato.aspirante.estado = 'BAJA'
                    contrato.aspirante.save()

                # F. Borrar contrato
                contrato.delete()

                # G. Refrescar el número oficial en la Base de Datos
                if cargo_obj:
                    cargo_obj.refrescar_conteo_plazas()

            # Usamos el PK del único movimiento creado
            pdf_url = reverse('imprimir_modelo_movimiento', kwargs={'pk': movimiento_baja.pk})

            return JsonResponse({
                'success': True, 
                'message': 'Contrato dado de baja y archivado correctamente.',
                'pdf_url': pdf_url
            })

        except Exception as e:
            # Captura cualquier error (Integridad, Modelo, etc.) y evita el crash del servidor
            print(f"ERROR AL DAR BAJA: {e}")
            return JsonResponse({
                'success': False, 
                'message': f'Error interno al procesar la baja: {str(e)}'
            }, status=500)
#*REPORTES

        


class MovimientoUpdateView(UpdateView):
    model = CAlta
    form_class = MovimientoForm
    template_name = "pages/contrato/movimiento_nomina.html"
    success_url = reverse_lazy('list_movimientos')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contrato = self.object
        context['contrato_actual'] = contrato
        context['aspirante'] = contrato.aspirante
        
        if contrato.cargo:
            # Usamos el método centralizado. ¡Ya no hay lógica duplicada!
            salario_base = contrato.calcular_salario_escala()
            context['salario_actual'] = salario_base # Pasa el número real al template

            context['initial_grupo'] = contrato.cargo.ncargo.grupo_escala
            context['initial_cat'] = contrato.cargo.ncargo.get_cat_ocupacional_display()
            context['initial_rol'] = contrato.cargo.rol.tipo if contrato.cargo.rol else "Cuadro"
            
            if salario_base:
                monto = float(salario_base)
                config = Configuracion.objects.first()
                fondo = float(config.fondo_tiempo_calc_tarif) if config and config.fondo_tiempo_calc_tarif else 190.6
                porcentaje_extra = float(config.porcentaje_horas_extras) if config and config.porcentaje_horas_extras else 0.25

                context['initial_salario_escala'] = round(monto, 2)
                context['initial_tarifa_horaria'] = round(monto / fondo, 5) if fondo else 0
                context['initial_tarifa_extras'] = round((context['initial_tarifa_horaria'] * porcentaje_extra) + context['initial_tarifa_horaria'], 5)
        
        # Añadimos la fecha límite cronológica para el JavaScript
        ultimo_mov = TMovimiento.objects.filter(
            aspirante=contrato.aspirante,
            no_expediente=contrato.no_expediente
        ).order_by('-fecha_efectiva').first()
        
        fecha_limite = ultimo_mov.fecha_efectiva if ultimo_mov else contrato.fecha_alta
        if fecha_limite:
            context['fecha_limite_js'] = fecha_limite.strftime('%Y-%m-%d')

        return context
    
    # ... (form_valid y form_invalid se quedan igual) ...
    @transaction.atomic
    def form_valid(self, form):
        try:
            # 1. Obtener contrato y fecha nueva
            contrato = self.object
            nueva_fecha = form.cleaned_data.get('fecha_efectiva')

            # =================================================================
            # PASO 1: VALIDACIÓN CRONOLÓGICA (EL PORTERO)
            # =================================================================
            
            # Buscamos si hay movimientos previos usando la ruta segura
            ultimo_mov = TMovimiento.objects.filter(
                aspirante=contrato.aspirante,
                no_expediente=contrato.no_expediente
            ).order_by('-fecha_efectiva').first()
            
            # La fecha límite es: La del último movimiento, O si no hay, la fecha de Alta original
            fecha_limite = ultimo_mov.fecha_efectiva if ultimo_mov else contrato.fecha_alta

            if nueva_fecha and nueva_fecha < fecha_limite:
                # Si la nueva fecha es viajar al pasado -> ERROR
                mensaje = f"Error Cronológico: La fecha seleccionada ({nueva_fecha.strftime('%d/%m/%Y')}) es anterior al último evento registrado ({fecha_limite.strftime('%d/%m/%Y')})."
                
                if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'form_is_valid': False, 'error_popup': mensaje}, status=400)
                else:
                    form.add_error('fecha_efectiva', mensaje)
                    return self.form_invalid(form)

            # =================================================================
            # PASO 1.5: VALIDACIÓN DE CAMBIOS REALES PARA MOVIMIENTO
            # =================================================================
            # Django sabe automáticamente si se modificó algún campo del formulario
            if not form.has_changed():
                mensaje = "No se ha aplicado ningún cambio válido. Modifique algún dato para generar un movimiento."
                
                if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'form_is_valid': False, 'error_popup': mensaje}, status=400)
                else:
                    form.add_error(None, mensaje)
                    return self.form_invalid(form)
            
            # Obtenemos el contrato previo seguro para extraer la historia
            contrato_previo = CAlta.objects.get(pk=self.object.pk)

            # =================================================================
            # PASO 2: CAPTURAR DATOS PREVIOS (LÓGICA DE NEGOCIO)
            # =================================================================
            cargo_ant = contrato_previo.cargo.ncargo.descripcion if contrato_previo.cargo else "---"
            unidad_ant = contrato_previo.cargo.departamento.unidad_organizativa.descripcion if (contrato_previo.cargo and contrato_previo.cargo.departamento) else "---"
            
            rol_ant_str = contrato_previo.cargo.rol.tipo if (contrato_previo.cargo and contrato_previo.cargo.rol) else "---"
            tri_ant_str = str(contrato_previo.tridente) if contrato_previo.tridente else "---"
            salario_ant = contrato_previo.calcular_salario_escala() or 0

            # --- NUEVO: CAPTURAR TIPO Y MOTIVO ANTERIOR ---
            tipo_ant_str = contrato_previo.tipo.descripcion if contrato_previo.tipo else "---"
            motivo_ant_str = contrato_previo.motivo.descripcion if contrato_previo.motivo else "---"

            # PASO 3: CAPTURAR DATOS NUEVOS
            cargo_nuevo_obj = form.cleaned_data.get('cargo')
            tipo_nuevo_obj = form.cleaned_data.get('tipo')
            motivo_nuevo_obj = form.cleaned_data.get('motivo')
            duracion_nueva_val = form.cleaned_data.get('duracion')
            
            cargo_nue = cargo_nuevo_obj.ncargo.descripcion if cargo_nuevo_obj else "---"
            unidad_nue = cargo_nuevo_obj.departamento.unidad_organizativa.descripcion if (cargo_nuevo_obj and cargo_nuevo_obj.departamento) else "---"
            
            salario_nue = float(self.request.POST.get('salarioEscala', 0))

            observaciones_txt = form.cleaned_data.get('observaciones', '')
            fecha_solicitud_dt = form.cleaned_data.get('fecha_solicitud')

            # PASO 4: GUARDAR EL HISTÓRICO
            from django.utils import timezone 
            
            # Extraemos la descripción del Tipo de Salario del formulario
            tipo_sal_nue_val = dict(form.fields['tipo_salario'].choices).get(form.cleaned_data.get('tipo_salario'), "---")

            nuevo_mov = TMovimiento.objects.create(
                contrato=self.object,
                aspirante=self.object.aspirante,
                no_expediente=self.object.no_expediente,
                fecha_efectiva=nueva_fecha if nueva_fecha else timezone.now().date(),
                fecha_solicitud=fecha_solicitud_dt,
                observaciones=observaciones_txt,
                
                # FOTO DEL PASADO
                cargo_anterior=cargo_ant,
                departamento_anterior=contrato_previo.cargo.departamento.descripcion if (contrato_previo.cargo and contrato_previo.cargo.departamento) else "---",
                grupo_escala_anterior=contrato_previo.cargo.ncargo.grupo_escala.nivel if (contrato_previo.cargo and contrato_previo.cargo.ncargo.grupo_escala) else "---",
                cat_ocupacional_anterior=contrato_previo.cargo.ncargo.get_cat_ocupacional_display() if contrato_previo.cargo else "---",
                tipo_salario_anterior=contrato_previo.get_tipo_salario_display() if contrato_previo.tipo_salario else "---",
                rol_anterior=rol_ant_str,       
                tridente_anterior=tri_ant_str,  
                salario_anterior=salario_ant,
                unidad_anterior=unidad_ant,
                tipo_contrato_anterior=tipo_ant_str, # <--- NUEVO
                motivo_anterior=motivo_ant_str,      # <--- NUEVO
                
                # FOTO DE LA NUEVA PROPUESTA
                cargo_nuevo=cargo_nue,
                departamento_nuevo=cargo_nuevo_obj.departamento.descripcion if (cargo_nuevo_obj and cargo_nuevo_obj.departamento) else "---",
                grupo_escala_nuevo=cargo_nuevo_obj.ncargo.grupo_escala.nivel if (cargo_nuevo_obj and cargo_nuevo_obj.ncargo.grupo_escala) else "---",
                cat_ocupacional_nuevo=cargo_nuevo_obj.ncargo.get_cat_ocupacional_display() if cargo_nuevo_obj else "---",
                tipo_salario_nuevo=tipo_sal_nue_val,
                rol_nuevo=cargo_nuevo_obj.rol.tipo if (cargo_nuevo_obj and cargo_nuevo_obj.rol) else "---",
                tridente_nuevo=str(form.cleaned_data.get('tridente')) if form.cleaned_data.get('tridente') else "---",
                salario_nuevo=salario_nue,
                unidad_nueva=unidad_nue,
                tipo_contrato_nuevo=tipo_nuevo_obj.descripcion if tipo_nuevo_obj else "---", # <--- NUEVO
                motivo_nuevo=motivo_nuevo_obj.descripcion if motivo_nuevo_obj else "---",   # <--- NUEVO
                duracion_nueva=duracion_nueva_val,                                           # <--- NUEVO
                
                tipo_movimiento="Movimiento de Nómina"
            )

            # PASO 5: ACTUALIZAR CONTRATO
            form.instance.en_proceso_movimiento = False
            self.object = form.save()

            messages.success(self.request, 'Movimiento de Nómina registrado correctamente.')

            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                

                return JsonResponse({
                    'form_is_valid': True, 
                    'message': 'Movimiento registrado correctamente.', 
                    'success_url': str(self.success_url),
                    # Ahora sí existe 'nuevo_mov'
                    'pdf_url': reverse('imprimir_modelo_movimiento', kwargs={'pk': nuevo_mov.pk})
                })
            return super().form_valid(form)

        
        except Exception as e:
            # --- MANEJO DE ERRORES ---
            print("\n" + "="*50)
            print("🔴 ERROR CRÍTICO EN MOVIMIENTO DE NÓMINA")
            print(f"Tipo: {type(e).__name__}")
            print(f"Mensaje: {str(e)}")
            print("-" * 20)
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
            print("="*50 + "\n")

            transaction.set_rollback(True)
            return JsonResponse({'form_is_valid': False, 'error_popup': str(e)}, status=500)

            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'form_is_valid': False,
                    'error_popup': f"Error del Sistema: {str(e)}",
                    'html_form': render_to_string(
                        self.template_name, 
                        self.get_context_data(form=form), 
                        request=self.request
                    )
                }, status=500)
            else:
                messages.error(self.request, f"Error crítico: {str(e)}")
                return self.form_invalid(form)

    def form_invalid(self, form):
        # --- 1. EL CHIVATO (Debug) ---
        # Esto imprimirá en tu terminal EXACTAMENTE por qué falla el formulario
        print("\n" + "!"*50)
        print("❌ ERROR DE VALIDACIÓN (400):")
        print(form.errors.as_json()) 
        print("!"*50 + "\n")
        
        # --- 2. RESPUESTA AL FRONTEND ---
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            # Renderizamos de nuevo el modal, ahora con los mensajes de error (rojos) que Django generó
            html = render_to_string(self.template_name, self.get_context_data(form=form), request=self.request)
            return JsonResponse({'form_is_valid': False, 'html_form': html}, status=400)
            
        return super().form_invalid(form)
    

def abreviar_cargo_inteligente(texto_cargo):
    if not texto_cargo: return "-"
    
    diccionario = {
        "OPERADOR": "OP.", "OPERARIO": "OPE.", "ESPECIALISTA": "ESP.",
        "MANTENIMIENTO": "MANT.", "DEPARTAMENTO": "DPTO.", "ADMINISTRATIVO": "ADMIN.",
        "ADMINISTRACION": "ADMIN.", "SERVICIOS": "SERVS.", "GENERAL": "GRAL.",
        "AUXILIAR": "AUX.", "TECNICO": "TEC.", "TÉCNICO": "TÉC.",
        "SUPERIOR": "SUP.", "PRINCIPAL": "PRAL.", "PRODUCCION": "PROD.",
        "FABRICACION": "FAB.", "FABRICACIÓN": "FAB.", "TRANSFORMADORES": "TRANSF.",
        "DISTRIBUCION": "DIST.", "ENERGETICO": "ENERG.", "MAQUINARIA": "MAQ.",
        "RECURSOS": "REC.", "HUMANOS": "HUM.", "SEGURIDAD": "SEG.",
    }
    
    palabras = texto_cargo.upper().split()
    # Pylance fix: Aseguramos que 'p' siempre es str y el resultado también
    palabras_nuevas = [str(diccionario.get(p, p)) for p in palabras]
    return " ".join(palabras_nuevas)


class ModeloMovimientoDocxView(View):
    def get(self, request, pk):
        print(f"[PDF] PK recibido: {pk}")
        
        
        # --- FASE 1: RESOLVER 404 ---
        obj = None
        es_movimiento_real = False

        # 1. Intentar como CAlta (pk de contrato)
        try:
            contrato = CAlta.objects.get(pk=pk)
            # Buscar el movimiento asociado a este contrato (el más reciente)
            movimiento = TMovimiento.objects.filter(contrato_id=contrato.pk).order_by('-id').first()
            if movimiento:
                obj = movimiento
                es_movimiento_real = True
            else:
                # Sin movimiento (no debería pasar, porque pdf_url vacío cuando no hay)
                obj = contrato
                es_movimiento_real = False
        except CAlta.DoesNotExist:
            # 2. Si no es contrato, intentar como TMovimiento directo
            try:
                obj = TMovimiento.objects.get(pk=pk)
                es_movimiento_real = True
            except TMovimiento.DoesNotExist:
                # 3. Fallback final a CBaja
                obj = CBaja.objects.filter(pk=pk).first()
                if not obj:
                    return HttpResponse("Error: No se encontró el registro.", status=404)
                es_movimiento_real = True
        print(f"[PDF] Objeto encontrado: {obj} (tipo: {type(obj)})")
        print(f"[PDF] Aspirante: {obj.aspirante.nombre} {obj.aspirante.papellido}")
        print(f"[PDF] fecha_solicitud: {getattr(obj, 'fecha_solicitud', 'NO_EXISTE')}")
        print(f"[PDF] observaciones: {getattr(obj, 'observaciones', 'NO_EXISTE')}")

        # Contrato base (CAlta activo). Si es baja, esto será None porque ya se borró.
        contrato_base = obj.contrato if es_movimiento_real and hasattr(obj, 'contrato') else (obj if not es_movimiento_real else None)

        # --- CONFIGURACIÓN TARIFA ---
        config = Configuracion.objects.first()
        fondo = float(config.fondo_tiempo_calc_tarif) if config and config.fondo_tiempo_calc_tarif else 190.6

        # --- FUNCIONES DE FORMATO LIBRES DE "NONE" ---
        def obtener_inicial_categoria(cat_str):
            if not cat_str or str(cat_str) == "None": return ""
            cat = str(cat_str).upper()
            if "OPERARIO" in cat: return "O"
            if "TÉCNICO" in cat or "TECNICO" in cat: return "T"
            if "CUADRO" in cat: return "C"
            if "SERVICIOS" in cat: return "S"
            if "ADMINISTRATIVO" in cat: return "A"
            return cat[0]

        def romano_a_arabigo(tri_obj):
            if not tri_obj or str(tri_obj) in ["None", "---"]: return ""
            t = str(tri_obj).upper()
            if "III" in t: return "3"
            if "II" in t: return "2"
            if "I" in t: return "1"
            return t

        # --- DETECCIÓN DE BAJA ---
        t_mov_str = str(getattr(obj, 'tipo_movimiento', "Alta Inicial" if not es_movimiento_real else "")).upper()
        es_alta = "ALTA" in t_mov_str
        es_baja = "BAJA" in t_mov_str
        es_movimiento = "MOVIMIENTO" in t_mov_str

        # Para las Bajas, necesitamos el registro histórico (CBaja) para sacar el Departamento y Categoría exactos
        baja_record = CBaja.objects.filter(no_expediente=obj.no_expediente).order_by('-id').first() if es_baja else None
        
        # El cargo de referencia (De donde sacamos dpto, ueb, categoria)
        cargo_ref = baja_record.cargo if (es_baja and baja_record) else (contrato_base.cargo if contrato_base else None)

        # --- LÓGICA DE DATOS SUPERIORES (UEB, Categoria, Cargo, Area) ---
        if es_baja:
            ueb_val = getattr(obj, 'unidad_anterior', "")
            if not ueb_val or ueb_val == "---":
                ueb_val = cargo_ref.departamento.unidad_organizativa.descripcion if cargo_ref else ""
            
            cargo_texto = getattr(obj, 'cargo_anterior', "")
            if not cargo_texto or cargo_texto == "---":
                cargo_texto = cargo_ref.ncargo.descripcion if cargo_ref else ""
                
        else:
            ueb_val = getattr(obj, 'unidad_nueva', "") if es_movimiento_real else ""
            if not ueb_val or ueb_val == "---":
                ueb_val = cargo_ref.departamento.unidad_organizativa.descripcion if cargo_ref else ""
            
            cargo_texto = getattr(obj, 'cargo_nuevo', "") if es_movimiento_real else ""
            if not cargo_texto or cargo_texto == "---":
                cargo_texto = cargo_ref.ncargo.descripcion if cargo_ref else ""

        # Área y Categoría
        area_val = getattr(obj, 'departamento_nuevo', None) if es_movimiento_real else ""
        if not area_val or area_val == "---":
            area_val = cargo_ref.departamento.descripcion if cargo_ref else ""

        cat_str = getattr(obj, 'cat_ocupacional_nuevo', None) if es_movimiento_real else ""
        if not cat_str or cat_str == "---":
            cat_str = cargo_ref.ncargo.get_cat_ocupacional_display() if cargo_ref and hasattr(cargo_ref.ncargo, 'get_cat_ocupacional_display') else ""
        
        cat_val = obtener_inicial_categoria(cat_str)

        # Extracción segura del Rol
        rol_n = getattr(obj, 'rol_nuevo', "") if es_movimiento_real else ""
        if not rol_n or rol_n == "---":
            rol_n = contrato_base.cargo.rol.tipo if contrato_base and contrato_base.cargo and getattr(contrato_base.cargo, 'rol', None) else ""

        # Ajuste de tamaño de texto para Cargo
        if len(cargo_texto) > 35:
            cargo_final = RichText(cargo_texto, size=14)
        elif len(cargo_texto) > 25:
            cargo_final = RichText(cargo_texto, size=16)
        else:
            cargo_final = cargo_texto

        # --- LÓGICA DE SALARIOS BASE ---
        salario_base_num = float(getattr(obj, 'salario_nuevo', 0) or 0) if es_movimiento_real else float(contrato_base.calcular_salario_escala() or 0)
        tarifa_base_num = salario_base_num / fondo if fondo else 0
        salario_str = f"{salario_base_num:.2f} ({tarifa_base_num:.5f})" if salario_base_num > 0 else ""

        # --- OTROS Y TOTAL ---
        suma_otros_num = 0
        if contrato_base:
            suma_otros_num = (
                float(getattr(contrato_base, 'maestria', 0) or 0) +
                float(getattr(contrato_base, 'doctorado', 0) or 0) +
                float(getattr(contrato_base, 'instructor', 0) or 0) +
                float(getattr(contrato_base, 'cnci', 0) or 0)
            )

        if suma_otros_num > 0:
            tarifa_otros_num = suma_otros_num / fondo if fondo else 0
            otros_str = f"{suma_otros_num:.2f} ({tarifa_otros_num:.5f})"
        else:
            otros_str = ""
        
        total_num = salario_base_num + suma_otros_num
        tarifa_total_num = total_num / fondo if fondo else 0
        total_str = f"{total_num:.2f} ({tarifa_total_num:.5f})" if total_num > 0 else ""

        # --- LÓGICA DE FECHAS (CORREGIDA) ---
        f_sol = getattr(obj, 'fecha_solicitud', None)
        fecha_efectiva = getattr(obj, 'fecha_efectiva', None) or getattr(obj, 'fecha_alta', datetime.now())

        # Fecha de solicitud (puede ser None)
        if f_sol:
            d_sol = f_sol.strftime("%d")
            m_sol = f_sol.strftime("%m")
            a_sol = f_sol.strftime("%Y")   # ← AÑO COMPLETO 4 DÍGITOS
        else:
            d_sol = ""
            m_sol = ""
            a_sol = ""

        # Fecha efectiva (siempre presente)
        ed = fecha_efectiva.strftime("%d")
        em = fecha_efectiva.strftime("%m")
        ea = fecha_efectiva.strftime("%Y")   # ← AÑO COMPLETO 4 DÍGITOS

        # --- CONSTRUCCIÓN DEL CONTEXTO PARA EL WORD ---
        context = {
            'doc_id': f"{obj.pk:06d}" if es_movimiento_real and not es_alta else "", 
            'ueb': ueb_val,
            
            # CABECERA SUPERIOR → FECHA DE SOLICITUD (o vacío)
            'd': d_sol,
            'm': m_sol,
            'a': a_sol,   # ← AÑO DE SOLICITUD (4 dígitos)
            
            # BLOQUE INFERIOR → FECHA EFECTIVA
            'ed': ed,
            'em': em,
            'ea': ea,     # ← AÑO DE EFECTIVA (4 dígitos)
            
            'x_alta': "X" if es_alta else "",
            'x_baja': "X" if es_baja else "",
            'x_mov':  "X" if es_movimiento else "",
            
            'nombre': obj.aspirante.nombre,
            'ap1': obj.aspirante.papellido,
            'ap2': obj.aspirante.sapellido,
            'exp': obj.no_expediente,
            
            'cat': cat_val,
            'cargo': cargo_final,
            'area': area_val,
            
            # --- FILA "A:" ---
            'rol_nue': "" if es_baja else rol_n,
            'tri_nue': "" if es_baja else romano_a_arabigo(getattr(obj, 'tridente_nuevo', contrato_base.tridente if contrato_base else "")),
            'sal_nue': "" if es_baja else salario_str,
            'otros': "" if es_baja else otros_str,
            'total_sal': "" if es_baja else total_str,
            
            'observaciones': getattr(obj, 'observaciones', "") or getattr(obj, 'motivo_movimiento', ""),
        }

        # --- FILA "DE:" (HISTÓRICO) ---
        if es_movimiento_real:
            sal_ant_num = float(getattr(obj, 'salario_anterior', 0) or 0)
            tarifa_ant_num = sal_ant_num / fondo if fondo else 0
            
            # Usamos los pagos extras del registro de baja si es baja, o del contrato actual si es mov
            base_ant = baja_record if es_baja else contrato_base
            suma_otros_ant = 0
            if base_ant:
                suma_otros_ant = (
                    float(getattr(base_ant, 'maestria', 0) or 0) + float(getattr(base_ant, 'doctorado', 0) or 0) +
                    float(getattr(base_ant, 'instructor', 0) or 0) + float(getattr(base_ant, 'cnci', 0) or 0)
                )

            if suma_otros_ant > 0:
                tarifa_otros_ant = suma_otros_ant / fondo if fondo else 0
                otros_ant_str = f"{suma_otros_ant:.2f} ({tarifa_otros_ant:.5f})"
            else:
                otros_ant_str = ""

            total_ant_num = sal_ant_num + suma_otros_ant
            tarifa_total_ant = total_ant_num / fondo if fondo else 0
            total_ant_str = f"{total_ant_num:.2f} ({tarifa_total_ant:.5f})" if total_ant_num > 0 else ""

            # Extraemos el Rol asegurándonos de que si viene la palabra "None" la borre
            rol_ant_bd = str(getattr(obj, 'rol_anterior', ""))
            
            context.update({
                'rol_ant': "" if rol_ant_bd == "None" else rol_ant_bd,
                'tri_ant': romano_a_arabigo(getattr(obj, 'tridente_anterior', "")),
                'sal_ant': f"{sal_ant_num:.2f} ({tarifa_ant_num:.5f})" if sal_ant_num > 0 else "",
                'otros_ant': otros_ant_str,
                'total_ant': total_ant_str,
            })

        # --- GENERACIÓN DEL DOCUMENTO ---
        try:
            template_path = os.path.join(settings.BASE_DIR, 'templates', 'pages', 'reportes', '13-MOVIMIENTO DE NOMINAS.docx')
            doc = DocxTemplate(template_path)
            doc.render(context)
            
            buffer = BytesIO()
            doc.save(buffer)
            buffer.seek(0)
            
            filename = f"Movimiento_{obj.no_expediente}.docx"
            response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return HttpResponse(f"Error al generar el documento: {str(e)}", status=500)
# 2. ACTUALIZA ESTA FUNCIÓN (Agregamos el retorno de textos para OOB)
# contratos/views.py

def cargar_salario(request):
    cargo_id      = request.GET.get('cargo')
    tridente_id   = request.GET.get('tridente')
    tipo_salario  = request.GET.get('tipo_salario')
    es_movimiento = request.GET.get('es_movimiento') 

    context = {
        'salario': 0.00, 'tarifa': 0, 'extras': 0,
        'nuevo_rol': '', 'nuevo_grupo': '', 
        'mostrar_tridente': False, 
        'es_movimiento': es_movimiento == '1',
        'es_cuadro_flag': False
    }

    if cargo_id:
        try:
            # 1. Cargamos el cargo con sus relaciones
            cargo = CargoPlantilla.objects.select_related('ncargo').get(id=cargo_id)
            grupo_obj = cargo.ncargo.grupo_escala
            rol_obj = cargo.rol  
            
            # 2. DETERMINAMOS SI ES CUADRO
            cat_nombre = cargo.ncargo.get_cat_ocupacional_display()
            categorias_cuadro = ["Cuadro Ejecutivo", "Cuadro Directivo", "CEJ", "CDI"]
            es_cuadro_por_cat = cat_nombre in categorias_cuadro or cargo.ncargo.cat_ocupacional in categorias_cuadro
            
            context['es_cuadro_flag'] = es_cuadro_por_cat
            context['nueva_cat'] = cat_nombre
            # AQUÍ LA CORRECCIÓN: Si no hay rol, se queda en "-", sin inventar "Cuadro"
            context['nuevo_rol'] = rol_obj.tipo if rol_obj else "-"
            context['nuevo_grupo'] = grupo_obj.nivel if grupo_obj else "-"
            
            monto = 0.00

            # --- ESCENARIO A: SALARIO FIJO DEL CARGO ---
            if tipo_salario == 'FIJ':
                monto = float(cargo.ncargo.salario_basico)
                context['mostrar_tridente'] = False
            
            # --- ESCENARIO B: SALARIO DINÁMICO DE LA TABLA ---
            else:
                salario_obj = None
                
                if es_cuadro_por_cat:
                    context['mostrar_tridente'] = False
                    salario_obj = NSalario.objects.filter(
                        grupo_escala=grupo_obj,
                        rol__isnull=True,
                        tridente__isnull=True
                    ).first()
                else:
                    nivel_num = nivel_romano_a_int(grupo_obj.nivel) if grupo_obj else 0
                    
                    if 22 <= nivel_num <= 24 and (rol_obj and rol_obj.tipo != "Decisorio"):
                        context['mostrar_tridente'] = False
                        salario_obj = None 
                    else:
                        context['mostrar_tridente'] = True 
                        
                        if tridente_id and rol_obj: 
                            salario_obj = NSalario.objects.filter(
                                grupo_escala=grupo_obj,
                                rol=rol_obj,
                                tridente_id=tridente_id
                            ).first()
                
                if salario_obj and salario_obj.monto:
                    monto = float(salario_obj.monto)

            # --- CÁLCULOS FINALES ---
            if monto > 0:
                config = Configuracion.objects.first()
                fondo = float(config.fondo_tiempo_calc_tarif) if config and config.fondo_tiempo_calc_tarif else 190.6
                
                context.update({
                    'salario': round(monto, 2),
                    'tarifa':  round(monto / fondo, 5) if fondo else 0,
                    'extras':  round(monto / 160.6, 5),
                })
                
        except Exception as e:
            print(f"Error calculando salario: {e}")

    return render(request, "pages/contrato/partials/cargar_salario.html", context)

# --- AÑADIR AL FINAL DE views.py ---

def cargar_cargos_contrato(request):
    """Vista exclusiva para rellenar el select de cargos en modales de Contratos vía HTMX"""
    id_dpto = request.GET.get('departamento')
    if id_dpto and id_dpto.isdigit():
        cargos = CargoPlantilla.objects.filter(
            departamento_id=id_dpto, 
            activo=True
        ).select_related('ncargo').order_by('ncargo__descripcion')
    else:
        cargos = CargoPlantilla.objects.none()
        
    # Renderiza estrictamente nuestra plantilla con los atributos data-
    return render(request, 'pages/contrato/partials/options_cargos.html', {'cargos': cargos})


# =========================================================================
# VISTA WIZARD FASE 3: Finalizar Contrato y (Opcional) Movimiento
# =========================================================================
def finalizar_contrato_wizard(request, aspirante_id):
    borrador = request.session.get('contrato_borrador')
    
    if not borrador:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'form_is_valid': False, 'error_popup': 'No hay datos en proceso. Cierre y vuelva a intentar.'}, status=400)
        return redirect('list_aspir')

    aspirante = get_object_or_404(Aspirante, doc_identidad=aspirante_id)

    # Convertir formato de diccionario para la creación dinámica del Modelo 
    # y filtrar campos extras del formulario que no existen en la BD (ej. unidad, departamento)
    borrador_db = borrador.copy()
    
    # Obtenemos los nombres válidos de los campos reales del modelo CAlta
    campos_validos = {f.name for f in CAlta._meta.get_fields()}
    atributos_validos = {getattr(f, 'attname', f.name) for f in CAlta._meta.get_fields()}

    for k in list(borrador_db.keys()):
        # Convertir FKs (ej. cargo -> cargo_id)
        if k + '_id' in atributos_validos and not k.endswith('_id'):
            borrador_db[k + '_id'] = borrador_db.pop(k)
        # Si el campo es del formulario pero NO del modelo, lo descartamos
        elif k not in campos_validos and k not in atributos_validos:
            borrador_db.pop(k)

    # GET: Construir el modal vacío (La columna izquierda oculta)
    if request.method == 'GET':
        try:
            unsaved_contrato = CAlta(**borrador_db)
            unsaved_contrato.aspirante = aspirante
        except Exception as e:
            print(f"🔴 ERROR AL INSTANCIAR CALTA: {e}")
            return JsonResponse({'form_is_valid': False, 'error_popup': f"Error interno en los datos: {str(e)}"}, status=500)
            
        # 1. PRECARGAR EL FORMULARIO CON LA INSTANCIA
        initial_data = {}
        if borrador_db.get('fecha_alta'):
            initial_data['fecha_efectiva'] = borrador_db.get('fecha_alta') 
            
        # --- INGENIERÍA INVERSA: Recuperar combos perdidos en la purga ---
        unidad_id = None
        dpto_id = None
        
        if unsaved_contrato.cargo_id:
            try:
                cargo_obj = CargoPlantilla.objects.select_related('departamento').get(pk=unsaved_contrato.cargo_id)
                if cargo_obj.departamento:
                    dpto_id = cargo_obj.departamento.id
                    unidad_id = cargo_obj.departamento.unidad_organizativa_id
                    
                    # Inyectamos al formulario para que los combos aparezcan seleccionados
                    initial_data['unidad'] = unidad_id
                    initial_data['departamento'] = dpto_id
                    initial_data['cargo'] = unsaved_contrato.cargo_id
            except CargoPlantilla.DoesNotExist:
                pass
                
        form_movimiento = MovimientoForm(instance=unsaved_contrato, initial=initial_data, user=request.user)
        
        # FORZAR QUERYSETS DE COMBOS DEPENDIENTES
        if unidad_id:
            form_movimiento.fields['departamento'].queryset = Departamento.objects.filter(unidad_organizativa_id=unidad_id)
        if dpto_id:
            form_movimiento.fields['cargo'].queryset = CargoPlantilla.objects.filter(departamento_id=dpto_id)

        # 2. CÁLCULO DE DATOS VISUALES (Columna Derecha)
        i_grupo = '-'
        i_cat = '-'
        i_rol = 'Cuadro'
        i_salario = 0
        i_tarifa = 0
        i_extras = 0

        if unsaved_contrato.cargo_id:
            try:
                cargo_obj = CargoPlantilla.objects.select_related('ncargo', 'rol').get(pk=unsaved_contrato.cargo_id)
                i_grupo = cargo_obj.ncargo.grupo_escala.nivel if cargo_obj.ncargo.grupo_escala else '-'
                i_cat = getattr(cargo_obj.ncargo, 'get_cat_ocupacional_display')() if hasattr(cargo_obj.ncargo, 'get_cat_ocupacional_display') else '-'
                i_rol = cargo_obj.rol.tipo if cargo_obj.rol else "Cuadro"
                
                
                # Rescatar cálculos salariales exactos
                monto = 0
                if unsaved_contrato.tipo_salario == 'FIJ':
                    monto = float(cargo_obj.ncargo.salario_basico)
                elif hasattr(unsaved_contrato, 'tridente_id') and unsaved_contrato.tridente_id:
                    sal_obj = NSalario.objects.filter(grupo_escala=cargo_obj.ncargo.grupo_escala, rol=cargo_obj.rol, tridente_id=unsaved_contrato.tridente_id).first()
                    if sal_obj: monto = float(sal_obj.monto)
                elif not cargo_obj.rol or cargo_obj.rol.tipo == "Cuadro":
                    sal_obj = NSalario.objects.filter(grupo_escala=cargo_obj.ncargo.grupo_escala, rol=cargo_obj.rol).first()
                    if sal_obj: monto = float(sal_obj.monto)

                if monto > 0:
                    config = Configuracion.objects.first()
                    fondo = float(config.fondo_tiempo_calc_tarif) if config and config.fondo_tiempo_calc_tarif else 190.6
                    i_salario = round(monto, 2)
                    i_tarifa = round(monto / fondo, 5) if fondo else 0
                    i_extras = round((i_tarifa*0.25)+i_tarifa, 5)
            except Exception as e:
                print(f"Error precargando datos del cargo: {e}")

        context = {
            'is_wizard': True,
            'aspirante': aspirante,
            'contrato_actual': unsaved_contrato, 
            'form': form_movimiento,
            'initial_rol': i_rol, 
            'initial_grupo': i_grupo, 
            'initial_cat': i_cat,
            'salario_actual': 0, # En wizard la izquierda SIEMPRE es 0 o vacía
            'initial_salario_escala': i_salario,
            'initial_tarifa_horaria': i_tarifa, 
            'initial_tarifa_extras': i_extras,
        }

        # =========================================================
        # LA MAGIA DEL WIZARD: Inyectar la fecha límite a la Fase 3
        # =========================================================
        
        
        ultima_baja = CBaja.objects.filter(aspirante_id=aspirante.id).order_by('-fecha_baja').first()
        ultimo_mov = TMovimiento.objects.filter(aspirante_id=aspirante.id).order_by('-fecha_efectiva').first()
        
        fechas_pasadas = []
        if ultima_baja and ultima_baja.fecha_baja: 
            fechas_pasadas.append(ultima_baja.fecha_baja)
        if ultimo_mov and ultimo_mov.fecha_efectiva: 
            fechas_pasadas.append(ultimo_mov.fecha_efectiva)
        
        if fechas_pasadas:
            fecha_limite = max(fechas_pasadas)
            context['fecha_limite_js'] = fecha_limite.strftime('%Y-%m-%d')
        # =========================================================

        return render(request, "pages/contrato/movimiento_nomina.html", context)

    # POST: Ejecutar el guardado en Base de Datos
    elif request.method == 'POST':
        generar_movimiento = request.POST.get('generar_movimiento') == 'true'
        from django.urls import reverse
        from django.utils import timezone
       
        
        try:
            with transaction.atomic():
                # 1. CREAR CONTRATO (Siempre se hace)
                contrato = CAlta(**borrador_db)
                contrato.aspirante = aspirante
                
                if contrato.cargo_id:
                    cargo_obj = CargoPlantilla.objects.get(pk=contrato.cargo_id)
                    if cargo_obj.rol:
                        contrato.rol = cargo_obj.rol
                
                # --- LA CURA CONTRA FECHAS ERRÓNEAS EN RECONTRATACIÓN ---
                # IMPERATIVO: Siempre mandará la fecha validada de la UI, ignorando la caché del borrador
                fecha_ui = request.POST.get('fecha_efectiva')
                if fecha_ui:
                    from datetime import datetime
                    contrato.fecha_alta = datetime.strptime(fecha_ui, '%Y-%m-%d').date()
                elif not contrato.fecha_alta:
                    from django.utils import timezone
                    contrato.fecha_alta = timezone.now().date()

                contrato.save() # Guardado Físico
                print(f"[WIZARD] Contrato creado con pk: {contrato.pk} - Aspirante: {contrato.aspirante.nombre}")
                
                pdf_url = ""
                
                # 2. CREAR MOVIMIENTO (Solo si el Caso 2 fue Confirmado)
                if generar_movimiento:
                    form = MovimientoForm(request.POST, instance=contrato, user=request.user)
                    if form.is_valid():
                        nueva_fecha = form.cleaned_data.get('fecha_efectiva')
                        cargo_nuevo = form.cleaned_data.get('cargo')
                        salario_nue = float(request.POST.get('salarioEscala', 0))
                        
                        tipo_sal_nue_val = dict(form.fields['tipo_salario'].choices).get(form.cleaned_data.get('tipo_salario'), "---")
                        TMovimiento.objects.create(
                            contrato=contrato,
                            aspirante=aspirante,
                            no_expediente=contrato.no_expediente,
                            fecha_efectiva=nueva_fecha if nueva_fecha else timezone.now().date(),
                            fecha_solicitud=form.cleaned_data.get('fecha_solicitud'),
                            observaciones=form.cleaned_data.get('observaciones', ''),
                            
                            cargo_anterior="---",
                            salario_anterior=0,
                            unidad_anterior="---",
                            
                            cargo_nuevo=cargo_nuevo.ncargo.descripcion if cargo_nuevo else "---",
                            departamento_nuevo=cargo_nuevo.departamento.descripcion if (cargo_nuevo and cargo_nuevo.departamento) else "---",
                            grupo_escala_nuevo=cargo_nuevo.ncargo.grupo_escala.nivel if (cargo_nuevo and cargo_nuevo.ncargo.grupo_escala) else "---",
                            cat_ocupacional_nuevo=cargo_nuevo.ncargo.get_cat_ocupacional_display() if cargo_nuevo else "---",
                            tipo_salario_nuevo=tipo_sal_nue_val,
                            rol_nuevo=cargo_nuevo.rol.tipo if (cargo_nuevo and cargo_nuevo.rol) else "---",
                            tridente_nuevo=str(form.cleaned_data.get('tridente')) if form.cleaned_data.get('tridente') else "---",
                            salario_nuevo=salario_nue,
                            unidad_nueva=cargo_nuevo.departamento.unidad_organizativa.descripcion if (cargo_nuevo and cargo_nuevo.departamento) else "---",
                            
                            tipo_movimiento="Alta Inicial"
                        )
                        form.instance.en_proceso_movimiento = False
                        contrato = form.save()
                        pdf_url = reverse('imprimir_modelo_movimiento', kwargs={'pk': contrato.pk})
                    else:
                        # Si el form falla, cancelamos todo (Contrato no se guarda) y devolvemos el error al modal
                        transaction.set_rollback(True)
                        context = {
                            'is_wizard': True, 'aspirante': aspirante, 'contrato_actual': contrato, 'form': form,
                            'initial_rol': '-', 'initial_grupo': '-', 'initial_cat': '-',
                            'salario_actual': 0, 'initial_salario_escala': 0,
                            'initial_tarifa_horaria': 0, 'initial_tarifa_extras': 0,
                        }
                        html = render_to_string("pages/contrato/movimiento_nomina.html", context, request=request)
                        return JsonResponse({'form_is_valid': False, 'html_form': html}, status=400)
                
                # 3. LIMPIEZA
                del request.session['contrato_borrador']
                if 'aspirante_borrador' in request.session:
                    del request.session['aspirante_borrador']
                
                return JsonResponse({
                    'form_is_valid': True, 
                    'success_url': reverse('list_contrato'),
                    'pdf_url': pdf_url
                })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'form_is_valid': False, 'error_popup': str(e)}, status=500)

def exportar_contratos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trabajadores Activos"

    # Estilos
    negrita_centrado = Font(bold=True, color="FFFFFF")
    fondo_azul = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alineacion_centro = Alignment(horizontal="center", vertical="center")

    # Cabeceras (Se añaden las tallas al final)
    cabeceras = [
        "No. Expediente", "Nombre", "1er Apellido", "2do Apellido",
        "Unidad Organizativa", "Departamento", "Cargo", "Cat. Ocupacional", "Tipo de Contrato", "Motivo",
        "Fecha Alta", "Duración (Dias)", "C. Formal", "Funcionario", "Designado",
        "Chofer Prof.", "Jubilado Recontratado", "Misión", "País Misión", "Reg. Militar",
        "Grupo Escala", "Rol", "Tridente", "Instructor", "CNCI",
        "Carnet de Ident.", "Edad", "Sexo", "Raza", "Nivel Educacional", "Especialidad", "Título de Oro",
        "Grado Científico", "Provincia", "Municipio", "Dirección", "Fijo", "Movil",
        "Talla Pantalón/Falda", "Talla Camisa/Blusa", "Talla Zapatos" # <--- NUEVAS COLUMNAS
    ]
    ws.append(cabeceras)

    # Formato cabeceras
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = negrita_centrado
        cell.fill = fondo_azul
        cell.alignment = alineacion_centro
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # Consulta Optimizada
    contratos = CAlta.objects.filter(aspirante__estado='ACTIVO').select_related(
        'aspirante', 'aspirante__municipio', 'aspirante__provincia', 'aspirante__especialidad',
        'tipo', 'motivo', 'cargo', 'cargo__ncargo', 'cargo__ncargo__grupo_escala',
        'cargo__departamento', 'cargo__departamento__unidad_organizativa', 'rol', 'tridente'
    ).order_by('cargo__departamento__unidad_organizativa__descripcion', 'aspirante__papellido')

    def si_no(valor): return "Sí" if valor else "No"

    for c in contratos:
        asp = c.aspirante

        # --- Obtener la categoría mapeada (más claro) ---
        cat_sigla = getattr(c.cargo.ncargo, 'cat_ocupacional', '') if c.cargo and c.cargo.ncargo else ''
        cat_valor = CAT_MAP.get(cat_sigla, cat_sigla[:1])  # Si no está en el mapa, usa la primera letra

        fila = [
            c.no_expediente, asp.nombre, asp.papellido, asp.sapellido,
            getattr(c.cargo.departamento.unidad_organizativa, 'descripcion', '') if c.cargo and c.cargo.departamento else "",
            getattr(c.cargo.departamento, 'descripcion', '') if c.cargo else "",
            getattr(c.cargo.ncargo, 'descripcion', '') if c.cargo else "",
            cat_valor,   # <--- AHORA SÍ SE USA LA VARIABLE
            getattr(c.tipo, 'descripcion', ''), getattr(c.motivo, 'descripcion', ''),
            c.fecha_alta.strftime('%d/%m/%Y') if c.fecha_alta else "", c.duracion or "",
            si_no(c.c_formal), si_no(c.funcionario), si_no(c.designado),
            si_no(c.profesional), si_no(c.jubilado_recontratado), si_no(c.mision), c.pais or "", c.reg_militar or "",
            # Grupo Escala
            getattr(c.cargo.ncargo.grupo_escala, 'nivel', '') if c.cargo and c.cargo.ncargo and getattr(c.cargo.ncargo, 'grupo_escala', None) else "",
            getattr(c.rol, 'tipo', ''), getattr(c.tridente, 'tipo', ''),
            c.instructor or 0, c.cnci or 0,
            asp.doc_identidad, asp.get_edad or "", asp.get_sexo_display(), asp.get_raza_display(),
            asp.get_nivel_educ_display(), getattr(asp.especialidad, 'nombre', ''), si_no(asp.titulo_oro),
            asp.grado_cientifico or "", getattr(asp.provincia, 'nombre', ''), getattr(asp.municipio, 'nombre', ''),
            asp.direccion or "", asp.movil_personal or "", asp.fijo_personal or "",
            # Tallas
            asp.tpantalon or "",
            asp.tcamisa or "",
            asp.tzapatos or ""
        ]
        ws.append(fila)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Listado_Activos_{datetime.now().strftime("%d_%m_%Y")}.xlsx"'
    wb.save(response)
    return response

class ExportarContratoWordView(FormView):
    template_name = "pages/contrato/exportar_contrato.html" # HTML que crearemos luego
    form_class = ExportarContratoWordForm

    def _get_contrato_optimizado(self, pk):
        """ Consulta unificada para evitar N+1 en las relaciones del contrato """
        try:
            return CAlta.objects.select_related(
                'aspirante__municipio',
                'aspirante__provincia',
                'aspirante__especialidad',
                'cargo__departamento__unidad_organizativa',
                'cargo__ncargo__grupo_escala',
                'cargo__rol',
                'tipo',
                'motivo',
                'tridente',
                'cla',
                'nocturnidad_1__nocturnidad',
                'nocturnidad_2__nocturnidad'
            ).get(pk=pk)
        except CAlta.DoesNotExist:
            raise Http404("El contrato solicitado no existe.")

    def get_form_kwargs(self):
        """Inyecta el queryset de cuadros activos al formulario."""
        kwargs = super().get_form_kwargs()

        from bolsa.models import Aspirante

        kwargs['cuadros_qs'] = (
            Aspirante.objects.filter(
                estado='ACTIVO',
                calta_contratos__cargo__ncargo__cat_ocupacional__in=['CDI', 'CEJ']
            )
            .distinct()
            .order_by('nombre', 'papellido')
        )

        return kwargs

    def get_initial(self):
        """ Inicializa el formulario transitorio con datos pre-procesados """
        initial = super().get_initial()
        contrato = self._get_contrato_optimizado(self.kwargs['pk'])
        
        config = Configuracion.objects.first()
        
        if config and config.periodo:
            initial['periodo_pago'] = 'Quincenal' if config.periodo == 15 else 'Mensual'
            
        if contrato.fecha_alta:
            initial['fecha_efectiva'] = contrato.fecha_alta

        # Extraer textos de nocturnidad
        textos_noc = []
        for noc_relation in [contrato.nocturnidad_1, contrato.nocturnidad_2]:
            if noc_relation and hasattr(noc_relation, 'nocturnidad') and noc_relation.nocturnidad:
                noc = noc_relation.nocturnidad
                if noc.hora_inicio and noc.hora_fin:
                    t_inicio = noc.hora_inicio.strftime('%I:%M %p').lstrip('0')
                    t_fin = noc.hora_fin.strftime('%I:%M %p').lstrip('0')
                    textos_noc.append(f"{t_inicio} a {t_fin}")

        if textos_noc:
            initial['nocturnidad_texto'] = " y ".join(textos_noc)

        return initial

    def get_context_data(self, **kwargs):
        """ Renderiza la pantalla (Estudio de Exportación) con datos de solo lectura para el usuario """
        context = super().get_context_data(**kwargs)
        contrato = self._get_contrato_optimizado(self.kwargs['pk'])
        context['contrato'] = contrato
        context['config'] = Configuracion.objects.first()
        context['provincias'] = NProvincia.objects.all().order_by('nombre')

        # Preparar Grados Científicos para solo lectura
        grados = []
        if getattr(contrato, 'maestria', 0) > 0:
            grados.append(f"Maestría (${contrato.maestria})")
        if getattr(contrato, 'doctorado', 0) > 0:
            grados.append(f"Doctorado (${contrato.doctorado})")
        context['grados_cientificos_texto'] = " • ".join(grados) if grados else "(-)"

        # Cálculos Matemáticos de Salario reutilizando el método del modelo
        monto_escala = contrato.calcular_salario_escala()
        context['salario_escala'] = f"{monto_escala:.2f}" if monto_escala else "0.00"
        context['tarifa_horaria'] = "0.00"
        
        config = Configuracion.objects.first()
        if monto_escala and config and config.fondo_tiempo_calc_tarif:
            fondo = float(config.fondo_tiempo_calc_tarif)
            tarifa = round(float(monto_escala) / fondo, 5) if fondo else 0
            context['tarifa_horaria'] = f"{tarifa:.5f}"

        return context

    def _extraer_fecha_nacimiento(self, doc_identidad):
        """ Extrae la fecha de nacimiento (dd/mm/yyyy) del Carnet de Identidad Cubano """
        if not doc_identidad or len(doc_identidad) < 6 or not doc_identidad[:6].isdigit():
            return "________"
        try:
            yy, mm, dd = int(doc_identidad[:2]), int(doc_identidad[2:4]), int(doc_identidad[4:6])
            hoy = datetime.today()
            century = 2000 if yy <= hoy.year % 100 else 1900
            fecha_nac = datetime(year=century + yy, month=mm, day=dd)
            return fecha_nac.strftime('%d/%m/%Y')
        except ValueError:
            return "________"

    def _compilar_contexto_word(self, contrato, cleaned_data, config):
        """ Consolida todos los datos en un diccionario listo para DocxTpl """
        
        # 1. Identificación Cuadro Firmante
        empleador_nombre = ""
        empleador_cargo = ""
        cuadro_aspirante = cleaned_data.get('firma_cuadro')
        
        if cuadro_aspirante:
            empleador_nombre = (
                f"{cuadro_aspirante.nombre} "
                f"{cuadro_aspirante.papellido} "
                f"{cuadro_aspirante.sapellido or ''}"
            ).strip()

            contrato_actual = (
                cuadro_aspirante.calta_contratos
                .filter(cargo__isnull=False)
                .select_related('cargo', 'cargo__ncargo')
                .order_by('-fecha_alta')
                .first()
            )

            if contrato_actual and contrato_actual.cargo and contrato_actual.cargo.ncargo:
                empleador_cargo = contrato_actual.cargo.ncargo.descripcion

        # 2. Fechas para el cierre del documento
        fecha_efectiva = cleaned_data.get('fecha_efectiva')
        if fecha_efectiva:
            firma_dia = str(fecha_efectiva.day).zfill(2)
            meses = {1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio', 
                     7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'}
            firma_mes = meses.get(fecha_efectiva.month, '').capitalize()
            firma_anno = str(fecha_efectiva.year)
        else:
            firma_dia = firma_mes = firma_anno = ""

        # 3. Matemática Salarial
        monto_escala = contrato.calcular_salario_escala()
        salario_escala = f"{monto_escala:.2f}" if monto_escala else ""
        tarifa_horaria = ""
        
        if monto_escala and config and config.fondo_tiempo_calc_tarif:
            fondo = float(config.fondo_tiempo_calc_tarif)
            tarifa_horaria = f"{round(float(monto_escala) / fondo, 5):.5f}" if fondo else ""

        # 4. Extracción de Provincia y Municipio (Desde el HTML manual)
        prov_id = self.request.POST.get('provincia_entidad')
        muni_id = self.request.POST.get('municipio_entidad')
        
        entidad_provincia = ""
        entidad_municipio = ""
        
        if prov_id:
            from nomencladores.models import NProvincia
            prov_obj = NProvincia.objects.filter(id=prov_id).first()
            if prov_obj: entidad_provincia = prov_obj.nombre
            
        if muni_id:
            from nomencladores.models import NMunicipio
            muni_obj = NMunicipio.objects.filter(id=muni_id).first()
            if muni_obj: entidad_municipio = muni_obj.nombre

        # 5. Resolución de Nombramiento (Limpiando el "Campo Vacío")
        res_no = ""
        if contrato.funcionario and contrato.funcionario_res:
            res_no = contrato.funcionario_res
        elif contrato.designado and contrato.designado_res:
            res_no = contrato.designado_res

        # 6. Construcción Final del Diccionario (Valores por defecto en "" para usar el Subrayado de Word)
        ctx = {
            # ENTIDAD Y FIRMANTE
            'entidad_nombre': (
                contrato.cargo.departamento.unidad_organizativa.descripcion
                if contrato.cargo and contrato.cargo.departamento and contrato.cargo.departamento.unidad_organizativa
                else ""
            ),
            'entidad_direccion': config.direccion if config else "",
            'entidad_rama': config.rama if config else "",
            'entidad_provincia': entidad_provincia, 
            'entidad_municipio': entidad_municipio, 
            'empleador_nombre': empleador_nombre,
            'empleador_cargo': empleador_cargo,

            # TRABAJADOR
            'trabajador_nombre': f"{contrato.aspirante.nombre} {contrato.aspirante.papellido} {contrato.aspirante.sapellido or ''}".strip(),
            'trabajador_ci': contrato.aspirante.doc_identidad,
            'trabajador_fecha_nacimiento': self._extraer_fecha_nacimiento(contrato.aspirante.doc_identidad).replace('_', ''),
            'trabajador_profesion': contrato.aspirante.especialidad.nombre if contrato.aspirante.especialidad else "",
            'trabajador_direccion': contrato.aspirante.direccion or "",
            'trabajador_provincia': contrato.aspirante.provincia.nombre if contrato.aspirante.provincia else "",
            'trabajador_municipio': contrato.aspirante.municipio.nombre if contrato.aspirante.municipio else "",

            # CONTRATO (CLÁUSULA PRIMERA Y SEGUNDA)
            'check_indeterminado': "X" if contrato.tipo and "INDETERMINADO" in contrato.tipo.descripcion.upper() else "",
            'check_determinado': "X" if contrato.tipo and "INDETERMINADO" not in contrato.tipo.descripcion.upper() else "",
            'contrato_duracion': f"{contrato.duracion} días." if contrato.duracion else "",
            'contrato_motivo': contrato.motivo.descripcion if contrato.motivo else "",
            'cargo_nombre': contrato.cargo.ncargo.descripcion if contrato.cargo else "",
            'cargo_grupo': contrato.cargo.ncargo.grupo_escala.nivel if contrato.cargo and contrato.cargo.ncargo.grupo_escala else "",
            
            # CATEGORÍA OCUPACIONAL
            'bg_obr': "X" if contrato.cargo and contrato.cargo.ncargo.cat_ocupacional == 'OPE' else "",
            'bg_adm': "X" if contrato.cargo and contrato.cargo.ncargo.cat_ocupacional == 'ADM' else "",
            'bg_ser': "X" if contrato.cargo and contrato.cargo.ncargo.cat_ocupacional == 'SER' else "",
            'bg_tec': "X" if contrato.cargo and contrato.cargo.ncargo.cat_ocupacional == 'TEC' else "",
            'bg_cdi': "X" if contrato.cargo and contrato.cargo.ncargo.cat_ocupacional in ['CDI', 'CEJ'] else "",

            # ROLES
            'check_fundamental': "X" if contrato.rol and 'FUNDAMENTAL' in contrato.rol.tipo.upper() else "",
            'check_apoyo': "X" if contrato.rol and 'APOYO' in contrato.rol.tipo.upper() else "",
            'check_decisorio': "X" if contrato.rol and 'DECISORIO' in contrato.rol.tipo.upper() else "",
            
            # TRIDENTES
            'check_t1': "X" if contrato.tridente and getattr(contrato.tridente, 'id', None) == 1 else "",
            'check_t2': "X" if contrato.tridente and getattr(contrato.tridente, 'id', None) == 2 else "",
            'check_t3': "X" if contrato.tridente and getattr(contrato.tridente, 'id', None) == 3 else "",

            # SALARIO Y HORARIOS
            'contrato_jornada': cleaned_data.get('jornada_texto') or "",
            'contrato_horario': cleaned_data.get('horario_texto') or "",
            'salario_resolucion': cleaned_data.get('resolucion_salario_escala') or "",
            'salario_escala': salario_escala,
            'tarifa_horaria': tarifa_horaria,
            'importe_total': salario_escala,
            
            # ADICIONALES Y PAGOS
            'pago_cla': f"(${contrato.cla.tarifa_hora})" if contrato.cla else "",
            'pago_nocturnidad': cleaned_data.get('nocturnidad_texto') or "",
            'pago_cnci': getattr(contrato, 'cnci', ""),
            'check_maestria': "X" if getattr(contrato, 'maestria', 0) > 0 else "",
            'maestria_importe': f"${contrato.maestria}" if getattr(contrato, 'maestria', 0) > 0 else "",
            'check_doctorado': "X" if getattr(contrato, 'doctorado', 0) > 0 else "",
            'doctorado_importe': f"${contrato.doctorado}" if getattr(contrato, 'doctorado', 0) > 0 else "",
            'sistema_pago_texto': cleaned_data.get('sistema_pago') or "",
            'check_pago_quincenal': "X" if cleaned_data.get('periodo_pago') == 'Quincenal' else "",
            'check_pago_mensual': "X" if cleaned_data.get('periodo_pago') == 'Mensual' else "",
            
            # RESOLUCIONES Y CIERRE
            'resolucion_nombramiento_no': res_no,
            'check_es_funcionario': "X" if contrato.funcionario else "",
            'check_es_designado': "X" if contrato.designado else "",
            'municipio_firma': entidad_municipio,
            'firma_dia': firma_dia,
            'firma_mes': firma_mes,
            'firma_anno': firma_anno,
        }
        return ctx

    def form_valid(self, form):
        """ Generación y Descarga del Documento Word Binario """
        contrato = self._get_contrato_optimizado(self.kwargs['pk'])
        
        config = Configuracion.objects.first()
        
        # Compilar Contexto (Inyectamos form.cleaned_data)
        context_word = self._compilar_contexto_word(contrato, form.cleaned_data, config)
        
        # Localizar plantilla
        template_path = os.path.join(settings.BASE_DIR, 'plantillas_word', 'documentos_legales', 'CONTRATO DE TRABAJO 2026.docx')
        
        if not os.path.exists(template_path):
            # Si el archivo fue borrado del disco, lanzamos el 404
            raise Http404("La plantilla legal oficial no se encuentra en el servidor. Ruta buscada: " + template_path)

        # Inyección a través de DocxTemplate
        doc = DocxTemplate(template_path)
        doc.render(context_word)
        
        # Preparación de la respuesta HTTP para descarga
        buffer = BytesIO()
        doc.save(buffer)
        buffer.seek(0)
        
        filename = f"Contrato_Trabajo_{contrato.aspirante.nombre}_{contrato.aspirante.papellido}.docx"
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response