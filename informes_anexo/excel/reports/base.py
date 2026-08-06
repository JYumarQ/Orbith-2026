import io
import openpyxl
from openpyxl.styles import Border, Side
from django.http import HttpResponse
from ..styles import Styles
from ..helpers import ExcelHelpers

class BaseExcelReport:
    """Clase base para todos los anexos de Orbith."""
    
    title = "INFORME INSTITUCIONAL"
    filename = "reporte.xlsx"
    
    def __init__(self, empresa_nombre="EMPRESA ELÉCTRICA CAMAGÜEY", reup="104-0-9082"):
        self.wb = openpyxl.Workbook()
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
        self.empresa_nombre = empresa_nombre
        self.reup = reup

    def generate(self, datos):
        raise NotImplementedError("Debes implementar generate() en la subclase.")

    def build_common_header(self, ws, title_text, unidad_nombre):
        # Título principal
        ExcelHelpers.merge_and_style(
            ws, 1, 1, 6, 1, 
            title_text, 
            font=Styles.FONT_TITLE, 
            fill=Styles.FILL_BRAND_DARK, 
            alignment=Styles.ALIGN_CENTER
        )
        
        # Metadatos institucionales
        unidad_str = str(unidad_nombre or "").upper()
        ws.cell(row=3, column=1, value=f"Empresa: {self.empresa_nombre}").font = Styles.FONT_HEADER
        ws.cell(row=3, column=5, value=f"Código REUP: {self.reup}").font = Styles.FONT_HEADER
        ws.cell(row=4, column=1, value="Organismo: MINEM").font = Styles.FONT_HEADER
        ws.cell(row=4, column=5, value="Provincia: Camagüey").font = Styles.FONT_HEADER
        ws.cell(row=5, column=1, value=f"Unidad: {unidad_str}").font = Styles.FONT_HEADER
        
        return 7  # Retorna la fila donde deben iniciar los encabezados de tabla

    def build_signature_block(self, ws, current_row):
        current_row += 2
        ws.cell(row=current_row, column=2, value="_________________________").alignment = Styles.ALIGN_CENTER
        ws.cell(row=current_row, column=5, value="_________________________").alignment = Styles.ALIGN_CENTER
        current_row += 1
        ws.cell(row=current_row, column=2, value="Director Capital Humano").alignment = Styles.ALIGN_CENTER
        ws.cell(row=current_row, column=5, value="Director General").alignment = Styles.ALIGN_CENTER
        return current_row

    def build_custom_signatures(self, ws, current_row, cargo_left="Director Capital Humano", nombre_left="", cargo_right="Director General", nombre_right=""):
        """
        Construye un bloque de firmas limpio centrado bajo la tabla de 6 columnas.
        Col 1-3: Firma Izquierda (Capital Humano)
        Col 4-6: Firma Derecha (Director General)
        """
        current_row += 3  # Separación visual respecto a la tabla o totales
        
        border_linea = Border(bottom=Side(style="thin"))
        
        # 1. Nombre del directivo en la celda superior
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        ws.cell(row=current_row, column=1, value=nombre_left or "").font = Styles.FONT_HEADER
        ws.cell(row=current_row, column=1).alignment = Styles.ALIGN_CENTER

        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=4, value=nombre_right or "").font = Styles.FONT_HEADER
        ws.cell(row=current_row, column=4).alignment = Styles.ALIGN_CENTER

        # Aplicar el borde inferior a TODAS las celdas del rango combinado para formar la línea continua
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = border_linea
        for col in range(4, 7):
            ws.cell(row=current_row, column=col).border = border_linea

        # 2. Cargo del directivo (Inmediatamente debajo de la línea)
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        ws.cell(row=current_row, column=1, value=cargo_left or "").font = Styles.FONT_DATA
        ws.cell(row=current_row, column=1).alignment = Styles.ALIGN_CENTER

        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=4, value=cargo_right or "").font = Styles.FONT_DATA
        ws.cell(row=current_row, column=4).alignment = Styles.ALIGN_CENTER

        return current_row

    def export_to_response(self, datos):
        self.generate(datos)
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        return response