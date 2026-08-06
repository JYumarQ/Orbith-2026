from openpyxl.utils import get_column_letter

class ExcelHelpers:
    @staticmethod
    def merge_and_style(ws, start_col, start_row, end_col, end_row, value, font=None, fill=None, alignment=None):
        """Fusiona celdas y aplica el formato en un solo paso."""
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=start_col, value=value)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        return cell

    @staticmethod
    def auto_fit_columns(ws, data_rows, default_widths=None):
        """Ajusta anchos basándose solo en celdas de datos para evitar desproporciones."""
        if not default_widths:
            default_widths = {1: 35, 2: 12, 3: 12, 4: 15, 5: 25, 6: 12, 7: 15}
            
        for row_idx in data_rows:
            for col_idx in range(1, 8):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    length = len(str(cell_value)) + 2
                    if length > default_widths.get(col_idx, 10):
                        default_widths[col_idx] = length

        for col_idx, width in default_widths.items():
            max_width = 50 if col_idx == 1 else 30
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width, max_width)