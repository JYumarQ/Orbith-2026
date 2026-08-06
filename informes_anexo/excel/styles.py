from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class Styles:
    COLOR_BRAND_DARK = "0F172A"
    COLOR_BRAND_LIGHT = "EEF2FF"
    COLOR_BORDER = "CBD5E1"
    
    FONT_TITLE = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    FONT_HEADER = Font(name='Arial', size=10, bold=True)
    FONT_DEPT = Font(name='Arial', size=10, bold=True, color="312E81")
    FONT_DATA = Font(name='Arial', size=10)
    FONT_TOTAL = Font(name='Arial', size=10, bold=True)
    
    FILL_BRAND_DARK = PatternFill(start_color=COLOR_BRAND_DARK, end_color=COLOR_BRAND_DARK, fill_type="solid")
    FILL_BRAND_LIGHT = PatternFill(start_color=COLOR_BRAND_LIGHT, end_color=COLOR_BRAND_LIGHT, fill_type="solid")
    
    _thin_side = Side(style='thin', color=COLOR_BORDER)
    _double_side = Side(style='double', color="000000")
    
    BORDER_ALL_THIN = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)
    BORDER_BOTTOM_THIN = Border(bottom=_thin_side)
    BORDER_TOTALS = Border(top=_thin_side, bottom=_double_side)
    
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
    ALIGN_INDENT = Alignment(horizontal='left', vertical='center', indent=2)