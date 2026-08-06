"""Pruebas de la reconciliación de hallazgos.

Son las primeras pruebas del proyecto, y están puestas donde un fallo sería silencioso
y caro: la reconciliación es la única pieza del módulo cuyo comportamiento no se puede
verificar leyendo el código, porque depende del estado acumulado en la base de datos a
lo largo de varios análisis.

Se ejecutan con:

    python manage.py test subsanacion

No hace falta instalar nada: es el runner que trae Django.
"""

from decimal import Decimal
from unittest.mock import patch

from django.contrib.contenttypes.models import ContentType
from django.test import TestCase
from django.utils import timezone

from .constantes import CRITICIDAD_ALTA, MODULO_CONTRATOS
from .contexto import ContextoAnalisis
from .models import EjecucionAnalisis, Hallazgo, RevisionHallazgo
from .reglas.base import HallazgoDetectado, ReglaBase
from .servicios import reconciliar_regla


class ReglaDePrueba(ReglaBase):
    """Regla de laboratorio: devuelve exactamente los hallazgos que se le pasan.

    `abstracta = True` es imprescindible: evita que se registre en el catálogo global.
    Sin eso, una regla de pruebas acabaría ejecutándose en los análisis de producción.
    `reconciliar_regla()` no consulta el registro —recibe la instancia— así que la
    regla funciona igual sin estar registrada.
    """

    abstracta = True

    codigo = 'PRUEBA-001'
    nombre = 'Regla de prueba'
    modulo = MODULO_CONTRATOS
    criticidad = CRITICIDAD_ALTA
    # Se usa un nomenclador sencillo solo para tener un ContentType válido. Los
    # `object_id` son inventados a propósito: el módulo nunca resuelve la
    # GenericForeignKey, y eso es justamente lo que le permite sobrevivir a que el
    # registro apuntado se borre.
    modelo = 'nomencladores.NProvincia'
    descripcion = 'Comprueba la reconciliación.'
    causa_probable = 'La provoca la propia prueba.'
    impacto = 'Ninguno: es una prueba.'
    solucion = 'No aplica.'

    def __init__(self, hallazgos=None):
        self.hallazgos = list(hallazgos or [])

    def ejecutar(self, contexto):
        yield from self.hallazgos


def hallazgo(object_id, valor='original'):
    """Un hallazgo detectado, con `datos` controlados para poder mover la huella."""
    return HallazgoDetectado(
        object_id=str(object_id),
        titulo=f'Hallazgo de prueba {object_id}',
        detalle='Detalle de prueba.',
        datos={'valor': valor},
    )


class ReconciliacionTests(TestCase):

    def setUp(self):
        self.ejecucion = EjecucionAnalisis.objects.create(
            estado=EjecucionAnalisis.COMPLETA)
        self.tipo = ContentType.objects.get_for_model(
            ReglaDePrueba().obtener_modelo())

    def _analizar(self, hallazgos):
        """Simula un barrido de esta regla y devuelve los contadores."""
        regla = ReglaDePrueba(hallazgos)
        return reconciliar_regla(regla, self.ejecucion, ContextoAnalisis(self.ejecucion))

    # ------------------------------------------------------------------
    def test_segundo_analisis_no_duplica_hallazgos(self):
        """La clave natural hace UPSERT: re-analizar no crea filas nuevas."""
        primera = self._analizar([hallazgo(1), hallazgo(2)])
        self.assertEqual(primera['nuevos'], 2)
        self.assertEqual(Hallazgo.objects.count(), 2)

        segunda = self._analizar([hallazgo(1), hallazgo(2)])
        self.assertEqual(segunda['nuevos'], 0)
        self.assertEqual(segunda['persistentes'], 2)
        self.assertEqual(segunda['resueltos'], 0)
        self.assertEqual(Hallazgo.objects.count(), 2)

        # Y se anota que se volvió a ver, que es lo que da idea de antigüedad real.
        self.assertEqual(
            list(Hallazgo.objects.values_list('veces_detectado', flat=True)), [2, 2])

    def test_hallazgo_ignorado_sobrevive_al_reanalisis(self):
        """Un «Ignorado» conserva su estado: el usuario ya dictaminó sobre él."""
        self._analizar([hallazgo(1)])
        ficha = Hallazgo.objects.get(object_id='1')
        ficha.estado = Hallazgo.IGNORADO
        ficha.nota = 'No aplica por una resolución especial.'
        ficha.save(update_fields=['estado', 'nota'])

        self._analizar([hallazgo(1)])

        ficha.refresh_from_db()
        self.assertEqual(ficha.estado, Hallazgo.IGNORADO)
        self.assertEqual(ficha.nota, 'No aplica por una resolución especial.')
        self.assertTrue(ficha.vigente)

    def test_hallazgo_que_desaparece_se_marca_corregido_y_se_archiva(self):
        """Si deja de detectarse, se archiva con su entrada en el historial.

        No se borra nunca: la ficha es la prueba de que el dato se corrigió.
        """
        self._analizar([hallazgo(1), hallazgo(2)])

        contadores = self._analizar([hallazgo(1)])
        self.assertEqual(contadores['resueltos'], 1)
        # La fila sigue existiendo.
        self.assertEqual(Hallazgo.objects.count(), 2)

        resuelto = Hallazgo.objects.get(object_id='2')
        self.assertFalse(resuelto.vigente)
        self.assertEqual(resuelto.estado, Hallazgo.CORREGIDO)
        self.assertTrue(resuelto.resuelto_automaticamente)
        self.assertIsNotNone(resuelto.fecha_resolucion)

        revision = resuelto.revisiones.get()
        self.assertTrue(revision.automatica)
        self.assertEqual(revision.estado_anterior, Hallazgo.PENDIENTE)
        self.assertEqual(revision.estado_nuevo, Hallazgo.CORREGIDO)
        self.assertIsNone(revision.usuario)

        # El que sigue detectándose no se toca.
        vigente = Hallazgo.objects.get(object_id='1')
        self.assertTrue(vigente.vigente)
        self.assertEqual(vigente.estado, Hallazgo.PENDIENTE)

    def test_ignorado_vuelve_a_pendiente_si_cambian_los_datos(self):
        """Si la huella cambia, ya no es el mismo caso y deja de estar silenciado.

        Es el mecanismo que impide que «Ignorado» se convierta en una alfombra bajo la
        que esconder problemas nuevos.
        """
        self._analizar([hallazgo(1, valor='original')])
        ficha = Hallazgo.objects.get(object_id='1')
        ficha.estado = Hallazgo.IGNORADO
        ficha.save(update_fields=['estado'])
        huella_original = ficha.huella

        self._analizar([hallazgo(1, valor='distinto')])

        ficha.refresh_from_db()
        self.assertNotEqual(ficha.huella, huella_original)
        self.assertEqual(ficha.estado, Hallazgo.PENDIENTE)
        self.assertIsNone(ficha.revisado_por)

        revision = ficha.revisiones.get()
        self.assertTrue(revision.automatica)
        self.assertEqual(revision.estado_anterior, Hallazgo.IGNORADO)
        self.assertEqual(revision.estado_nuevo, Hallazgo.PENDIENTE)
        self.assertIn('cambiaron', revision.nota)

    def test_corregido_que_reaparece_vuelve_a_pendiente(self):
        """Si vuelve a detectarse, no estaba corregido de verdad."""
        self._analizar([hallazgo(1)])
        self._analizar([])                      # desaparece -> Corregido
        self.assertEqual(Hallazgo.objects.get(object_id='1').estado, Hallazgo.CORREGIDO)

        self._analizar([hallazgo(1)])           # reaparece

        ficha = Hallazgo.objects.get(object_id='1')
        self.assertEqual(ficha.estado, Hallazgo.PENDIENTE)
        self.assertTrue(ficha.vigente)
        self.assertIsNone(ficha.fecha_resolucion)
        self.assertFalse(ficha.resuelto_automaticamente)

    def test_una_regla_truncada_no_marca_nada_como_corregido(self):
        """Si la regla se corta por el límite, no se puede saber qué desapareció.

        Marcar como corregidos los que no se alcanzaron sería mentir, y además es la
        propiedad que hace correcto ejecutar el análisis por partes.
        """
        self._analizar([hallazgo(1), hallazgo(2), hallazgo(3)])

        with patch('subsanacion.servicios.LIMITE_HALLAZGOS_POR_REGLA', 1):
            contadores = self._analizar([hallazgo(1), hallazgo(2), hallazgo(3)])

        self.assertTrue(contadores['truncada'])
        self.assertEqual(contadores['resueltos'], 0)
        self.assertEqual(Hallazgo.objects.filter(vigente=True).count(), 3)

    def test_clave_extra_separa_hallazgos_del_mismo_registro(self):
        """Varios problemas sobre el mismo registro no se pisan entre sí."""
        uno = HallazgoDetectado(
            object_id='1', clave_extra='campo_a', titulo='Falta el campo A',
            datos={'campo': 'A'})
        dos = HallazgoDetectado(
            object_id='1', clave_extra='campo_b', titulo='Falta el campo B',
            datos={'campo': 'B'})

        self._analizar([uno, dos])
        self.assertEqual(Hallazgo.objects.filter(object_id='1').count(), 2)

        # Se puede ignorar uno y dejar el otro pendiente.
        ficha_a = Hallazgo.objects.get(object_id='1', clave_extra='campo_a')
        ficha_a.estado = Hallazgo.IGNORADO
        ficha_a.save(update_fields=['estado'])

        self._analizar([uno, dos])

        self.assertEqual(
            Hallazgo.objects.get(object_id='1', clave_extra='campo_a').estado,
            Hallazgo.IGNORADO)
        self.assertEqual(
            Hallazgo.objects.get(object_id='1', clave_extra='campo_b').estado,
            Hallazgo.PENDIENTE)

    def test_hallazgo_global_no_se_duplica(self):
        """Un hallazgo global (object_id vacío) también respeta la clave natural.

        En PostgreSQL `NULL != NULL`, así que si `content_type` fuera nulable la
        restricción única no impediría duplicados y cada barrido crearía otra fila. Por
        eso las reglas globales declaran modelo y usan object_id = ''.
        """
        global_ = HallazgoDetectado(
            object_id='', clave_extra='ausente', titulo='No existe la configuración')

        self._analizar([global_])
        self._analizar([global_])

        self.assertEqual(Hallazgo.objects.filter(object_id='').count(), 1)


class ExportacionTests(TestCase):
    """La exportación tiene que contener exactamente lo que hay en pantalla.

    Es el invariante que hace fiable el módulo: un Excel descargado con filtros puestos
    que trajera más o menos filas que la pantalla llevaría a conclusiones falsas sobre
    cuántos problemas hay.
    """

    def setUp(self):
        from .constantes import CRITICIDAD_CRITICA, CRITICIDAD_MEDIA, MODULO_NOMINA

        tipo = ContentType.objects.get_for_model(ReglaDePrueba().obtener_modelo())
        ahora = timezone.now()
        self.datos = [
            ('EXP-001', MODULO_CONTRATOS, CRITICIDAD_CRITICA,
             'Pérez Gómez, Juan (85010112345) · contrato sin cargo'),
            ('EXP-002', MODULO_CONTRATOS, CRITICIDAD_MEDIA,
             'Ruiz Díaz, Ana (90020254321) · motivo incoherente'),
            ('EXP-003', MODULO_NOMINA, CRITICIDAD_MEDIA,
             'Pérez Gómez, Juan (85010112345) · salario desincronizado'),
        ]
        for indice, (codigo, modulo, criticidad, titulo) in enumerate(self.datos, start=1):
            Hallazgo.objects.create(
                codigo_regla=codigo, content_type=tipo, object_id=str(indice),
                modulo=modulo, criticidad=criticidad, titulo=titulo,
                ultima_deteccion=ahora, estado=Hallazgo.PENDIENTE)

    def _filas_de_excel(self, contenido):
        import io

        from openpyxl import load_workbook

        hoja = load_workbook(io.BytesIO(contenido)).active
        # La fila 4 es la cabecera; los datos empiezan en la 5 y terminan antes del total.
        return [fila for fila in hoja.iter_rows(min_row=5, values_only=True)
                if fila[0] and not str(fila[0]).startswith('Total')]

    def _comparar(self, querystring):
        from django.test import RequestFactory

        from .exportar import exportar_a_excel
        from .filtros import filtrar_hallazgos

        peticion = RequestFactory().get(f'/subsanacion/exportar/excel/{querystring}')
        esperadas = filtrar_hallazgos(peticion)[0].count()
        obtenidas = len(self._filas_de_excel(exportar_a_excel(peticion).content))
        self.assertEqual(
            obtenidas, esperadas,
            f'Con «{querystring or "sin filtros"}» la pantalla muestra {esperadas} '
            f'hallazgos y el Excel trae {obtenidas}.')
        return esperadas

    def test_sin_filtros_exporta_todos_los_abiertos(self):
        self.assertEqual(self._comparar(''), 3)

    def test_respeta_el_filtro_de_gravedad(self):
        self.assertEqual(self._comparar('?solo_criticos=1'), 1)

    def test_respeta_el_filtro_de_categoria(self):
        from .constantes import MODULO_NOMINA
        self.assertEqual(self._comparar(f'?modulo={MODULO_NOMINA}'), 1)

    def test_respeta_la_busqueda(self):
        """Busca por carnet, que es lo que las reglas escriben en el título."""
        self.assertEqual(self._comparar('?q=85010112345'), 2)

    def test_los_archivados_no_salen_salvo_que_se_pidan(self):
        Hallazgo.objects.filter(codigo_regla='EXP-003').update(vigente=False)
        self.assertEqual(self._comparar(''), 2)
        self.assertEqual(self._comparar('?archivados=1'), 3)

    def test_el_pdf_se_genera_sin_errores(self):
        from django.test import RequestFactory

        from .exportar import exportar_a_pdf

        peticion = RequestFactory().get('/subsanacion/exportar/pdf/')
        respuesta = exportar_a_pdf(peticion)
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta['Content-Type'], 'application/pdf')
        self.assertTrue(respuesta.content.startswith(b'%PDF'))


class IndiceDeSaludTests(TestCase):
    """La fórmula del índice: porcentaje de registros limpios."""

    def test_sin_hallazgos_el_indice_es_cien(self):
        from .servicios import indice_desde_afectados
        self.assertEqual(indice_desde_afectados(0, 250), Decimal('100.00'))

    def test_todos_afectados_da_cero(self):
        from .servicios import indice_desde_afectados
        self.assertEqual(indice_desde_afectados(250, 250), Decimal('0.00'))

    def test_nunca_baja_de_cero_ni_se_satura(self):
        """La fórmula anterior se desbordaba; ésta no puede.

        Aunque haya más hallazgos que registros (un registro puede generar varios), el
        número de registros AFECTADOS nunca supera el de registros analizados, así que el
        índice se queda siempre entre 0 y 100 y sigue siendo informativo.
        """
        from .servicios import indice_desde_afectados
        self.assertEqual(indice_desde_afectados(9999, 100), Decimal('0.00'))
        self.assertEqual(indice_desde_afectados(35, 247), Decimal('85.83'))

    def test_sin_registros_no_divide_por_cero(self):
        from .servicios import indice_desde_afectados
        self.assertEqual(indice_desde_afectados(0, 0), Decimal('100.00'))


class NotaObligatoriaTests(TestCase):
    """Ignorar un hallazgo o marcarlo como falso positivo exige explicar por qué."""

    def setUp(self):
        tipo = ContentType.objects.get_for_model(ReglaDePrueba().obtener_modelo())
        self.hallazgo = Hallazgo.objects.create(
            codigo_regla='PRUEBA-NOTA', content_type=tipo, object_id='1',
            modulo=MODULO_CONTRATOS, criticidad=CRITICIDAD_ALTA,
            titulo='Hallazgo de prueba', ultima_deteccion=timezone.now(),
            estado=Hallazgo.PENDIENTE)

    def test_ignorar_sin_nota_se_rechaza(self):
        from .servicios import cambiar_estado_hallazgo

        with self.assertRaises(ValueError):
            cambiar_estado_hallazgo(self.hallazgo, Hallazgo.IGNORADO, nota='')

        self.hallazgo.refresh_from_db()
        self.assertEqual(self.hallazgo.estado, Hallazgo.PENDIENTE)

    def test_falso_positivo_sin_nota_se_rechaza(self):
        from .servicios import cambiar_estado_hallazgo

        with self.assertRaises(ValueError):
            cambiar_estado_hallazgo(self.hallazgo, Hallazgo.FALSO_POSITIVO, nota='   ')

    def test_ignorar_con_nota_se_acepta(self):
        from .servicios import cambiar_estado_hallazgo

        cambiar_estado_hallazgo(
            self.hallazgo, Hallazgo.IGNORADO, nota='No aplica en este caso.')
        self.hallazgo.refresh_from_db()
        self.assertEqual(self.hallazgo.estado, Hallazgo.IGNORADO)

    def test_marcar_en_revision_no_exige_nota(self):
        """Solo Ignorado y Falso positivo exigen nota; el resto de estados no."""
        from .servicios import cambiar_estado_hallazgo

        cambiar_estado_hallazgo(self.hallazgo, Hallazgo.EN_REVISION, nota='')
        self.hallazgo.refresh_from_db()
        self.assertEqual(self.hallazgo.estado, Hallazgo.EN_REVISION)


class AvisoDeCriticasTests(TestCase):
    """El receptor de `notificaciones` avisa a los administradores de una crítica nueva.

    Vive en `subsanacion/tests.py` y no en `notificaciones` porque lo que se prueba es
    el CONTRATO de la señal (qué hallazgos cuentan como «nuevos» y qué le llega al
    receptor), no los detalles internos de `notificaciones`. El receptor en sí está en
    `notificaciones/receptores.py`, conectado desde `notificaciones/apps.py::ready()`.
    """

    def setUp(self):
        from usuarios.models import CustomUser

        self.admin = CustomUser(username='admin_prueba', es_admin=True, is_active=True)
        self.admin.set_password('contraseña-de-prueba')
        self.admin.save()
        self.ejecucion = EjecucionAnalisis.objects.create(estado=EjecucionAnalisis.COMPLETA)
        self.tipo = ContentType.objects.get_for_model(ReglaDePrueba().obtener_modelo())

    def test_una_critica_nueva_notifica_a_los_admins(self):
        from notificaciones.models import Notificacion
        from subsanacion.constantes import CRITICIDAD_CRITICA
        from subsanacion.signals import analisis_finalizado

        Hallazgo.objects.create(
            codigo_regla='PRUEBA-CRIT', content_type=self.tipo, object_id='1',
            modulo=MODULO_CONTRATOS, criticidad=CRITICIDAD_CRITICA,
            titulo='Crítica de prueba', ultima_deteccion=timezone.now(),
            ejecucion=self.ejecucion, estado=Hallazgo.PENDIENTE)

        analisis_finalizado.send(sender=EjecucionAnalisis, ejecucion=self.ejecucion)

        self.assertEqual(
            Notificacion.objects.filter(destinatario=self.admin).count(), 1)

    def test_una_critica_antigua_no_notifica_de_nuevo(self):
        """Un hallazgo que ya existía antes de esta ejecución no es «nuevo»."""
        from notificaciones.models import Notificacion
        from subsanacion.constantes import CRITICIDAD_CRITICA
        from subsanacion.signals import analisis_finalizado

        antigua = timezone.now() - timezone.timedelta(days=10)
        hallazgo = Hallazgo.objects.create(
            codigo_regla='PRUEBA-VIEJA', content_type=self.tipo, object_id='2',
            modulo=MODULO_CONTRATOS, criticidad=CRITICIDAD_CRITICA,
            titulo='Crítica antigua', ultima_deteccion=timezone.now(),
            ejecucion=self.ejecucion, estado=Hallazgo.PENDIENTE)
        Hallazgo.objects.filter(pk=hallazgo.pk).update(primera_deteccion=antigua)

        analisis_finalizado.send(sender=EjecucionAnalisis, ejecucion=self.ejecucion)

        self.assertEqual(Notificacion.objects.filter(destinatario=self.admin).count(), 0)

    def test_una_critica_de_media_no_notifica(self):
        from notificaciones.models import Notificacion
        from subsanacion.constantes import CRITICIDAD_MEDIA
        from subsanacion.signals import analisis_finalizado

        Hallazgo.objects.create(
            codigo_regla='PRUEBA-MEDIA', content_type=self.tipo, object_id='3',
            modulo=MODULO_CONTRATOS, criticidad=CRITICIDAD_MEDIA,
            titulo='Media de prueba', ultima_deteccion=timezone.now(),
            ejecucion=self.ejecucion, estado=Hallazgo.PENDIENTE)

        analisis_finalizado.send(sender=EjecucionAnalisis, ejecucion=self.ejecucion)

        self.assertEqual(Notificacion.objects.filter(destinatario=self.admin).count(), 0)

    def test_admin_sin_unidades_ve_su_notificacion_personal(self):
        """El caso real que motivó `destinatario`: un admin sin unidades asignadas."""
        from notificaciones.views import _notificaciones_del_usuario
        from subsanacion.constantes import CRITICIDAD_CRITICA
        from subsanacion.signals import analisis_finalizado

        self.assertEqual(self.admin.unidades.count(), 0)

        Hallazgo.objects.create(
            codigo_regla='PRUEBA-VISIBLE', content_type=self.tipo, object_id='4',
            modulo=MODULO_CONTRATOS, criticidad=CRITICIDAD_CRITICA,
            titulo='Visible para el admin', ultima_deteccion=timezone.now(),
            ejecucion=self.ejecucion, estado=Hallazgo.PENDIENTE)
        analisis_finalizado.send(sender=EjecucionAnalisis, ejecucion=self.ejecucion)

        self.assertEqual(_notificaciones_del_usuario(self.admin).count(), 1)


class ReglasRetiradasTests(TestCase):
    """Los hallazgos de una regla que ya no existe no pueden quedarse abiertos.

    La reconciliación se acota por `codigo_regla`, así que si una regla se retira su
    consulta no vuelve a ejecutarse y sus hallazgos se quedarían vigentes para siempre,
    penalizando el índice sin que nada pudiera comprobarlos.
    """

    def setUp(self):
        from .constantes import CLAVES_MODULOS
        self.ejecucion = EjecucionAnalisis.objects.create(
            estado=EjecucionAnalisis.EN_CURSO,
            modulos_solicitados=list(CLAVES_MODULOS),
            modulos_completados=list(CLAVES_MODULOS),
        )
        tipo = ContentType.objects.get_for_model(ReglaDePrueba().obtener_modelo())
        self.huerfano = Hallazgo.objects.create(
            codigo_regla='RETIRADA-999', content_type=tipo, object_id='7',
            modulo=MODULO_CONTRATOS, criticidad=CRITICIDAD_ALTA,
            titulo='Hallazgo de una regla que ya no existe',
            ultima_deteccion=timezone.now(), estado=Hallazgo.PENDIENTE)

    def test_analisis_completo_archiva_los_huerfanos(self):
        from .servicios import archivar_hallazgos_de_reglas_retiradas

        archivados = archivar_hallazgos_de_reglas_retiradas(self.ejecucion)
        self.assertEqual(archivados, 1)

        self.huerfano.refresh_from_db()
        self.assertFalse(self.huerfano.vigente)
        # NO se marca como corregido: nadie lo corrigió, solo dejó de comprobarse.
        self.assertEqual(self.huerfano.estado, Hallazgo.PENDIENTE)
        self.assertFalse(self.huerfano.resuelto_automaticamente)

        revision = self.huerfano.revisiones.get()
        self.assertTrue(revision.automatica)
        self.assertIn('se retiró', revision.nota)

    def test_analisis_parcial_no_archiva_nada(self):
        """En un análisis parcial no se puede distinguir «retirada» de «no ejecutada»."""
        from .servicios import archivar_hallazgos_de_reglas_retiradas

        self.ejecucion.modulos_solicitados = [MODULO_CONTRATOS]
        self.ejecucion.modulos_completados = [MODULO_CONTRATOS]
        self.ejecucion.save()

        self.assertEqual(archivar_hallazgos_de_reglas_retiradas(self.ejecucion), 0)
        self.huerfano.refresh_from_db()
        self.assertTrue(self.huerfano.vigente)


class RegistroDeReglasTests(TestCase):
    """El catálogo real tiene que estar bien formado."""

    def test_las_reglas_declaran_todo_lo_obligatorio(self):
        from .reglas import obtener_reglas

        reglas = obtener_reglas()
        self.assertGreater(len(reglas), 0, 'No se descubrió ninguna regla.')

        codigos = [regla.codigo for regla in reglas]
        self.assertEqual(len(codigos), len(set(codigos)),
                         'Hay códigos de regla duplicados.')

        for regla in reglas:
            with self.subTest(regla=regla.codigo):
                # Los cuatro textos de la guía son la razón de ser del módulo: un
                # hallazgo sin explicación ni solución no ayuda a nadie.
                self.assertTrue(regla.descripcion)
                self.assertTrue(regla.causa_probable)
                self.assertTrue(regla.impacto)
                self.assertTrue(regla.solucion)
                # El modelo tiene que resolverse de verdad.
                self.assertIsNotNone(regla.obtener_modelo())

    def test_toda_regla_tiene_un_enlace_al_registro(self):
        """Si un hallazgo no lleva a ninguna parte, el usuario no puede corregirlo."""
        from .enlaces import ENLACES_POR_MODELO
        from .reglas import obtener_reglas

        for regla in obtener_reglas():
            with self.subTest(regla=regla.codigo):
                self.assertIn(
                    regla.modelo_del_enlace(), ENLACES_POR_MODELO,
                    f'La regla {regla.codigo} apunta a «{regla.modelo_del_enlace()}», '
                    f'que no está en ENLACES_POR_MODELO, así que sus hallazgos no '
                    f'tendrían botón para llegar al registro.')


class FuncionesPurasEnCalienteTests(TestCase):
    """Las funciones puras que comparten `ejecutar()` y `evaluar_instancia()`.

    Son la garantía real de que el barrido periódico y la alerta en caliente nunca
    puedan divergir: como ambos métodos llaman a la MISMA función, probarla una vez
    cubre a los dos. No hace falta un CAlta real: son funciones de valores escalares.
    """

    def test_grado_cientifico_sin_reflejar(self):
        from .reglas.contratos import _grado_cientifico_sin_reflejar

        self.assertEqual(_grado_cientifico_sin_reflejar('MC', 0, 0), [
            ('maestria_vacia', 'Máster', 'campo_maestria')])
        self.assertEqual(_grado_cientifico_sin_reflejar('DC', 0, 0), [
            ('doctorado_vacio', 'Doctor', 'campo_doctorado')])
        self.assertEqual(_grado_cientifico_sin_reflejar('MC', 30, 0), [])
        self.assertEqual(_grado_cientifico_sin_reflejar(None, 0, 0), [])

    def test_funcionario_o_designado_sin_codigo(self):
        from .reglas.contratos import _funcionario_o_designado_sin_codigo

        self.assertEqual(
            _funcionario_o_designado_sin_codigo(True, False, '', ''),
            [('funcionario_sin_codigo', 'Funcionario', 'funcionario_res')])
        self.assertEqual(
            _funcionario_o_designado_sin_codigo(False, True, '', None),
            [('designado_sin_codigo', 'Designado', 'designado_res')])
        self.assertEqual(
            _funcionario_o_designado_sin_codigo(True, True, 'RES-1', 'RES-2'), [])
        self.assertEqual(
            _funcionario_o_designado_sin_codigo(False, False, '', ''), [])

    def test_fechas_de_vencimiento_faltantes(self):
        from datetime import date

        from .reglas.contratos import _fechas_de_vencimiento_faltantes

        hoy = date.today()
        self.assertEqual(_fechas_de_vencimiento_faltantes(None, None, None), [
            ('fecha_vence_lic', 'licencia'),
            ('fecha_vence_recal', 'recalificación'),
            ('fecha_vence_seg', 'seguro')])
        self.assertEqual(_fechas_de_vencimiento_faltantes(hoy, hoy, hoy), [])
        self.assertEqual(_fechas_de_vencimiento_faltantes(hoy, None, hoy), [
            ('fecha_vence_recal', 'recalificación')])

    def test_campo_de_texto_vacio(self):
        from .reglas.aspirantes import _campo_de_texto_vacio

        self.assertTrue(_campo_de_texto_vacio(None))
        self.assertTrue(_campo_de_texto_vacio(''))
        self.assertTrue(_campo_de_texto_vacio('   '))
        self.assertFalse(_campo_de_texto_vacio('54212345'))

    def test_mision_sin_pais(self):
        from .reglas.contratos import _mision_sin_pais

        self.assertTrue(_mision_sin_pais(True, None))
        self.assertTrue(_mision_sin_pais(True, ''))
        self.assertFalse(_mision_sin_pais(True, 'Brasil'))
        self.assertFalse(_mision_sin_pais(False, None))

    def test_c_formal_sin_res(self):
        from .reglas.contratos import _c_formal_sin_res

        self.assertTrue(_c_formal_sin_res(True, None))
        self.assertTrue(_c_formal_sin_res(True, ''))
        self.assertFalse(_c_formal_sin_res(True, 'RES-1'))
        self.assertFalse(_c_formal_sin_res(False, None))


class EvaluarInstanciaTests(TestCase):
    """`evaluar_instancia()` de CTR-007/012/013 sobre una instancia real, no un dict.

    Se usa `SimpleNamespace` en vez de un `CAlta` real: los tres métodos solo acceden
    a atributos propios y a un único salto de FK ya resuelto (`contrato.aspirante`),
    exactamente la restricción O(1) que exige `ReglaBase.evaluar_instancia()` — así que
    un objeto con esos atributos puestos a mano ejercita el mismo código que un CAlta
    de verdad, sin necesitar el grafo completo de fixtures (aspirante, cargo,
    plantilla, etc.).
    """

    def test_grado_cientifico_detecta_master_sin_reflejar(self):
        from types import SimpleNamespace

        from .reglas.contratos import GradoCientificoSinReflejar

        contrato = SimpleNamespace(
            pk=1, maestria=0, doctorado=0,
            aspirante=SimpleNamespace(grado_cientifico='MC'))

        detectados = GradoCientificoSinReflejar().evaluar_instancia(contrato)

        self.assertEqual(len(detectados), 1)
        self.assertEqual(detectados[0].clave_extra, 'maestria_vacia')

    def test_grado_cientifico_sin_problema_devuelve_none(self):
        from types import SimpleNamespace

        from .reglas.contratos import GradoCientificoSinReflejar

        contrato = SimpleNamespace(
            pk=1, maestria=30, doctorado=0,
            aspirante=SimpleNamespace(grado_cientifico='MC'))

        self.assertIsNone(GradoCientificoSinReflejar().evaluar_instancia(contrato))

    def test_funcionario_sin_codigo_detectado(self):
        from types import SimpleNamespace

        from .reglas.contratos import FuncionarioODesignadoSinCodigo

        contrato = SimpleNamespace(
            pk=1, funcionario=True, designado=False,
            funcionario_res='', designado_res='')

        detectados = FuncionarioODesignadoSinCodigo().evaluar_instancia(contrato)

        self.assertEqual(len(detectados), 1)
        self.assertEqual(detectados[0].clave_extra, 'funcionario_sin_codigo')

    def test_chofer_sin_fechas_detectado(self):
        from types import SimpleNamespace

        from .reglas.contratos import ChoferSinFechasDeVencimiento

        contrato = SimpleNamespace(
            pk=1, profesional=True,
            fecha_vence_lic=None, fecha_vence_recal=None, fecha_vence_seg=None)

        detectados = ChoferSinFechasDeVencimiento().evaluar_instancia(contrato)

        self.assertEqual(len(detectados), 1)
        self.assertIn('licencia', detectados[0].detalle)

    def test_chofer_no_profesional_devuelve_none(self):
        from types import SimpleNamespace

        from .reglas.contratos import ChoferSinFechasDeVencimiento

        contrato = SimpleNamespace(
            pk=1, profesional=False,
            fecha_vence_lic=None, fecha_vence_recal=None, fecha_vence_seg=None)

        self.assertIsNone(ChoferSinFechasDeVencimiento().evaluar_instancia(contrato))

    def test_movil_personal_vacio_detectado(self):
        from types import SimpleNamespace

        from .reglas.aspirantes import MovilPersonalVacio

        aspirante = SimpleNamespace(pk=1, movil_personal='')

        detectados = MovilPersonalVacio().evaluar_instancia(aspirante)

        self.assertEqual(len(detectados), 1)

    def test_movil_personal_completo_devuelve_none(self):
        from types import SimpleNamespace

        from .reglas.aspirantes import MovilPersonalVacio

        aspirante = SimpleNamespace(pk=1, movil_personal='54212345')

        self.assertIsNone(MovilPersonalVacio().evaluar_instancia(aspirante))

    def test_direccion_vacia_detectada(self):
        from types import SimpleNamespace

        from .reglas.aspirantes import DireccionVacia

        aspirante = SimpleNamespace(pk=1, direccion=None)

        detectados = DireccionVacia().evaluar_instancia(aspirante)

        self.assertEqual(len(detectados), 1)

    def test_nivel_educativo_vacio_detectado(self):
        from types import SimpleNamespace

        from .reglas.aspirantes import NivelEducativoSinCompletar

        aspirante = SimpleNamespace(pk=1, nivel_educ=None)

        detectados = NivelEducativoSinCompletar().evaluar_instancia(aspirante)

        self.assertEqual(len(detectados), 1)

    def test_nivel_educativo_sin_acreditar_no_es_vacio(self):
        """'SA' (Sin Acreditar) es una respuesta válida del formulario, no un vacío."""
        from types import SimpleNamespace

        from .reglas.aspirantes import NivelEducativoSinCompletar

        aspirante = SimpleNamespace(pk=1, nivel_educ='SA')

        self.assertIsNone(NivelEducativoSinCompletar().evaluar_instancia(aspirante))

    def test_mision_sin_pais_detectada(self):
        from types import SimpleNamespace

        from .reglas.contratos import MisionSinPais

        contrato = SimpleNamespace(pk=1, mision=True, pais=None)

        detectados = MisionSinPais().evaluar_instancia(contrato)

        self.assertEqual(len(detectados), 1)

    def test_sin_mision_devuelve_none(self):
        from types import SimpleNamespace

        from .reglas.contratos import MisionSinPais

        contrato = SimpleNamespace(pk=1, mision=False, pais=None)

        self.assertIsNone(MisionSinPais().evaluar_instancia(contrato))

    def test_c_formal_sin_codigo_detectado(self):
        from types import SimpleNamespace

        from .reglas.contratos import ConformidadFormalSinCodigo

        contrato = SimpleNamespace(pk=1, c_formal=True, c_formal_res='')

        detectados = ConformidadFormalSinCodigo().evaluar_instancia(contrato)

        self.assertEqual(len(detectados), 1)

    def test_sin_c_formal_devuelve_none(self):
        from types import SimpleNamespace

        from .reglas.contratos import ConformidadFormalSinCodigo

        contrato = SimpleNamespace(pk=1, c_formal=False, c_formal_res='')

        self.assertIsNone(ConformidadFormalSinCodigo().evaluar_instancia(contrato))


class EvaluarReglasDeInstanciaTests(TestCase):
    """El despachador de `en_caliente.py`: solo reglas `evaluable_en_caliente` y del
    modelo correcto entran a evaluarse; solo las que detectan algo salen en el resultado."""

    def test_filtra_por_evaluable_en_caliente_y_modelo(self):
        from .en_caliente import evaluar_reglas_de_instancia

        candidata = ReglaDePrueba(hallazgos=[
            HallazgoDetectado(object_id='1', titulo='Detectado')])
        candidata.evaluable_en_caliente = True
        # `ReglaDePrueba` no implementa evaluar_instancia(); se le añade uno mínimo
        # que reutiliza sus propios hallazgos de prueba, para aislar el despachador.
        candidata.evaluar_instancia = lambda instancia: candidata.hallazgos

        no_evaluable = ReglaDePrueba(hallazgos=[
            HallazgoDetectado(object_id='1', titulo='No debería aparecer')])
        no_evaluable.evaluable_en_caliente = False

        from nomencladores.models import NProvincia

        from . import en_caliente

        with patch.object(en_caliente, 'obtener_reglas', return_value=[candidata, no_evaluable]):
            instancia = NProvincia.objects.create(nombre='Provincia de en_caliente')
            resultados = en_caliente.evaluar_reglas_de_instancia(instancia)

        self.assertEqual(len(resultados), 1)
        regla, detectados = resultados[0]
        self.assertIs(regla, candidata)
        self.assertEqual(detectados, candidata.hallazgos)
