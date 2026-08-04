"""Exportación del listado de hallazgos a Excel y PDF.

Las dos exportaciones y la pantalla comparten `filtros.filtrar_hallazgos()`, que es la
única forma de garantizar que lo que el usuario descarga es exactamente lo que está
viendo: mismos filtros, misma búsqueda, mismo orden. Es el mismo criterio que ya usa
`informes_anexo`, donde una sola función de datos alimenta el HTML, el Excel y el PDF.

Se exporta lo que hay en pantalla, no la tabla entera.

Las dos bibliotecas (`openpyxl` y `xhtml2pdf`) ya estaban en `requirements.txt`: no se
añade ninguna dependencia nueva.
"""

import logging
from io import BytesIO

from django.http import HttpResponse
from django.template.loader import get_template
from django.utils import timezone

from .constantes import ETIQUETA_MODULO
from .filtros import filtrar_hallazgos

logger = logging.getLogger(__name__)

# Sin `datos` ni `detalle` completos: el Excel es para trabajar en él (filtrar, repartir
# el trabajo, marcar avance), no para leer la explicación entera de cada hallazgo. Eso
# está en la pantalla de detalle.
COLUMNAS = (
    ('Código', 18),
    ('Gravedad', 14),
    ('Categoría', 22),
    ('Estado', 14),
    ('Hallazgo', 80),
    ('Registro afectado', 20),
    ('Primera detección', 18),
    ('Última detección', 18),
    ('Veces detectado', 15),
    ('Revisado por', 20),
    ('Observación', 40),
)


def _filas(consulta):
    """Convierte los hallazgos en filas planas, compartidas por Excel y PDF."""
    for hallazgo in consulta.select_related('content_type', 'revisado_por'):
        tipo = hallazgo.content_type
        yield {
            'codigo': hallazgo.codigo_regla,
            'gravedad': hallazgo.get_criticidad_display(),
            'criticidad': hallazgo.criticidad,
            'categoria': ETIQUETA_MODULO.get(hallazgo.modulo, hallazgo.modulo),
            'estado': hallazgo.get_estado_display(),
            'titulo': hallazgo.titulo,
            'detalle': hallazgo.detalle,
            'datos': hallazgo.datos or {},
            'registro': (
                f'{tipo.model} #{hallazgo.object_id}' if hallazgo.object_id
                else 'Hallazgo global'),
            'primera': hallazgo.primera_deteccion,
            'ultima': hallazgo.ultima_deteccion,
            'veces': hallazgo.veces_detectado,
            'revisado_por': (
                hallazgo.revisado_por.username if hallazgo.revisado_por else ''),
            'nota': hallazgo.nota or '',
        }


def _nombre_archivo(extension):
    return f'subsanacion_hallazgos_{timezone.localtime():%Y%m%d_%H%M}.{extension}'


def _resumen_filtros(filtros):
    """Describe en una frase los filtros aplicados, para que el archivo se explique solo.

    Sin esto, un Excel descargado con filtros puestos parece la lista completa y alguien
    puede concluir que hay menos problemas de los que hay.
    """
    if not filtros.get('hay_filtros'):
        return 'Todos los hallazgos pendientes y en revisión'

    partes = []
    if filtros.get('q'):
        partes.append(f'búsqueda «{filtros["q"]}»')
    if filtros.get('solo_criticos'):
        partes.append('solo críticos')
    if filtros.get('modulos'):
        partes.append('categorías: ' + ', '.join(
            ETIQUETA_MODULO.get(clave, clave) for clave in filtros['modulos']))
    if filtros.get('archivados'):
        partes.append('incluyendo archivados')
    return 'Filtrado por ' + '; '.join(partes) if partes else 'Con filtros aplicados'


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def exportar_a_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    consulta, filtros = filtrar_hallazgos(request)

    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Hallazgos'

    # Encabezado con el contexto del informe.
    hoja['A1'] = 'Subsanación · Hallazgos de integridad de datos'
    hoja['A1'].font = Font(size=14, bold=True, color='1E293B')
    hoja['A2'] = (f'Generado el {timezone.localtime():%d/%m/%Y a las %H:%M} · '
                  f'{_resumen_filtros(filtros)}')
    hoja['A2'].font = Font(size=9, italic=True, color='64748B')

    fila_cabecera = 4
    relleno = PatternFill('solid', fgColor='1E293B')
    for columna, (titulo, ancho) in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=fila_cabecera, column=columna, value=titulo)
        celda.font = Font(bold=True, color='FFFFFF', size=10)
        celda.fill = relleno
        celda.alignment = Alignment(vertical='center')
        hoja.column_dimensions[get_column_letter(columna)].width = ancho

    total = 0
    for indice, fila in enumerate(_filas(consulta), start=fila_cabecera + 1):
        total += 1
        hoja.cell(row=indice, column=1, value=fila['codigo'])
        hoja.cell(row=indice, column=2, value=fila['gravedad'])
        hoja.cell(row=indice, column=3, value=fila['categoria'])
        hoja.cell(row=indice, column=4, value=fila['estado'])
        hoja.cell(row=indice, column=5, value=fila['titulo'])
        hoja.cell(row=indice, column=6, value=fila['registro'])
        # Las fechas se escriben sin zona horaria: openpyxl no admite datetime con tzinfo.
        hoja.cell(row=indice, column=7,
                  value=timezone.localtime(fila['primera']).replace(tzinfo=None))
        hoja.cell(row=indice, column=8,
                  value=timezone.localtime(fila['ultima']).replace(tzinfo=None))
        hoja.cell(row=indice, column=9, value=fila['veces'])
        hoja.cell(row=indice, column=10, value=fila['revisado_por'])
        hoja.cell(row=indice, column=11, value=fila['nota'])

        for columna in (7, 8):
            hoja.cell(row=indice, column=columna).number_format = 'DD/MM/YYYY HH:MM'

    hoja.cell(row=fila_cabecera + total + 2, column=1,
              value=f'Total: {total} hallazgo{"s" if total != 1 else ""}').font = Font(bold=True)

    # Fijar la cabecera y activar el autofiltro para que la hoja sea usable de verdad.
    hoja.freeze_panes = hoja.cell(row=fila_cabecera + 1, column=1)
    if total:
        hoja.auto_filter.ref = (
            f'A{fila_cabecera}:{get_column_letter(len(COLUMNAS))}{fila_cabecera + total}')

    flujo = BytesIO()
    libro.save(flujo)
    flujo.seek(0)

    respuesta = HttpResponse(
        flujo.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    respuesta['Content-Disposition'] = f'attachment; filename="{_nombre_archivo("xlsx")}"'
    return respuesta


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def exportar_a_pdf(request):
    from xhtml2pdf import pisa

    consulta, filtros = filtrar_hallazgos(request)
    filas = list(_filas(consulta))

    plantilla = get_template('pages/subsanacion/partials/hallazgos_pdf.html')
    html = plantilla.render({
        'filas': filas,
        'total': len(filas),
        'resumen_filtros': _resumen_filtros(filtros),
        'generado': timezone.localtime(),
    })

    respuesta = HttpResponse(content_type='application/pdf')
    respuesta['Content-Disposition'] = f'attachment; filename="{_nombre_archivo("pdf")}"'

    resultado = pisa.pisaDocument(BytesIO(html.encode('UTF-8')), respuesta)
    if resultado.err:
        logger.error('Subsanación: falló la generación del PDF (%s errores).', resultado.err)
        return HttpResponse(
            'No se pudo generar el PDF. Inténtelo de nuevo o exporte a Excel.',
            status=500)
    return respuesta
